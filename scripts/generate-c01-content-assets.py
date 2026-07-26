#!/usr/bin/env python3
"""Generate C-01 specification data, variant prices, Section H data, and PDFs.

Inputs remain in the main clone as required by Event B. The script verifies the
approved workbook before reading it and refuses any divergence result that does
not match the governing L16 amendment.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import re
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from typing import Any

import openpyxl
from reportlab.lib import colors
from reportlab.lib.enums import TA_LEFT
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import (
    BaseDocTemplate,
    Frame,
    KeepTogether,
    PageTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
)

ROOT = Path(__file__).resolve().parents[1]
EXPECTED_BYTES = 679_988
EXPECTED_MD5 = "2bb681dff71ae744ea4d44418a09476a"

SIZE_ORDER = [
    "10x10",
    "20x8",
    "20x10",
    "20x12",
    "30x10",
    "40x8",
    "20x20",
    "40x10",
    "40x12",
]
MINI_SIZE_ORDER = SIZE_ORDER[:4]

PRODUCTS: dict[str, dict[str, str]] = {
    "porta-cabins": {
        "name": "Porta Cabins",
        "sheet": "01 Porta Cabins",
        "canonical": "https://www.samanportable.com/product/porta-cabins",
    },
    "low-cost-porta-cabin": {
        "name": "Low Cost Porta Cabin",
        "sheet": "03 Low Cost Porta Cabin",
        "canonical": "https://www.samanportable.com/product/porta-cabins/low-cost-porta-cabin",
    },
    "luxury-porta-cabin": {
        "name": "Luxury Porta Cabin",
        "sheet": "04 Luxury Porta Cabin",
        "canonical": "https://www.samanportable.com/product/porta-cabins/luxury-porta-cabin",
    },
    "mini-porta-cabin": {
        "name": "Mini Porta Cabin",
        "sheet": "05 Mini Porta Cabin",
        "canonical": "https://www.samanportable.com/product/porta-cabins/mini-porta-cabin",
    },
    "ms-porta-cabin": {
        "name": "MS Porta Cabin",
        "sheet": "06 MS Porta Cabin",
        "canonical": "https://www.samanportable.com/product/porta-cabins/ms-porta-cabin",
    },
    "porta-cabin-shop": {
        "name": "Porta Cabin Shop",
        "sheet": "08 Porta Cabin Shop",
        "canonical": "https://www.samanportable.com/product/porta-cabins/porta-cabin-shop",
    },
    "porta-cabin-with-toilet": {
        "name": "Porta Cabin with Toilet",
        "sheet": "09 Porta Cabin with Toilet",
        "canonical": "https://www.samanportable.com/product/porta-cabins/porta-cabin-with-toilet",
    },
    "portacabin-office": {
        "name": "Portacabin Office",
        "sheet": "10 Portacabin Office",
        "canonical": "https://www.samanportable.com/product/porta-cabins/portacabin-office",
    },
    "steel-porta-cabin": {
        "name": "Steel Porta Cabin",
        "sheet": "13 Steel Porta Cabin",
        "canonical": "https://www.samanportable.com/product/porta-cabins/steel-porta-cabin",
    },
}

SPEC_ROW_GROUPS = [
    ("Steel Structure", range(14, 22)),
    ("Walls, Roof, Floor & Insulation", range(25, 35)),
    ("Doors, Windows, Electrical & Services", range(38, 50)),
]

EXPECTED_DIVERGENCE = {
    "low-cost-porta-cabin": (10, "mini-porta-cabin", 3),
    "luxury-porta-cabin": (10, "portacabin-office", 3),
    "mini-porta-cabin": (11, "low-cost-porta-cabin", 3),
    "ms-porta-cabin": (10, "steel-porta-cabin", 5),
    "steel-porta-cabin": (10, "ms-porta-cabin", 5),
    "porta-cabin-shop": (13, "porta-cabin-with-toilet", 14),
    "porta-cabin-with-toilet": (14, "porta-cabin-shop", 14),
    "portacabin-office": (13, "luxury-porta-cabin", 3),
}

HARD_COMMON = {
    "Welding & fabrication",
    "Fasteners & sealing",
    "Grills / mosquito mesh",
    "Electrical wiring",
    "Electrical protection",
    "Warranty",
}

WARRANTY_SENTENCE = (
    "5-year structural warranty and 1-year finishing warranty as standard; "
    "finishing warranty extendable to 2 years on request, confirmed at quotation."
)


def round_rupee(value: Any) -> int:
    return int(Decimal(str(value)).quantize(Decimal("1"), rounding=ROUND_HALF_UP))


def indian(value: int) -> str:
    digits = str(abs(int(value)))
    if len(digits) <= 3:
        grouped = digits
    else:
        grouped = digits[-3:]
        head = digits[:-3]
        while head:
            grouped = head[-2:] + "," + grouped
            head = head[:-2]
    return ("-" if value < 0 else "") + grouped


def verify_workbook(path: Path) -> None:
    size = path.stat().st_size
    digest = hashlib.md5(path.read_bytes()).hexdigest()
    if size != EXPECTED_BYTES or digest != EXPECTED_MD5:
        raise SystemExit(
            f"Workbook verification failed: bytes={size}, md5={digest}; "
            f"expected bytes={EXPECTED_BYTES}, md5={EXPECTED_MD5}"
        )


def extract_spec_rows(workbook: openpyxl.Workbook, sheet_name: str) -> list[dict[str, Any]]:
    sheet = workbook[sheet_name]
    rows: list[dict[str, Any]] = []
    for group, indices in SPEC_ROW_GROUPS:
        for row_number in indices:
            component = sheet.cell(row_number, 1).value
            detail = sheet.cell(row_number, 4).value
            if not component or detail is None:
                raise SystemExit(f"Missing specification cell: {sheet_name}!row {row_number}")
            rows.append(
                {
                    "group": group,
                    "component": str(component),
                    "detail": str(detail),
                    "sourceSheet": sheet_name,
                    "sourceRow": row_number,
                }
            )
    if len(rows) != 30:
        raise SystemExit(f"{sheet_name}: expected 30 specification rows, found {len(rows)}")
    return rows


def set_detail(rows: list[dict[str, Any]], component: str, detail: str, source: str) -> None:
    for row in rows:
        if row["component"] == component:
            row["detail"] = detail
            row["selectionSource"] = source
            return
    raise SystemExit(f"Cannot map missing row {component!r}")


def apply_amendment(
    workbook: openpyxl.Workbook, specs: dict[str, list[dict[str, Any]]]
) -> None:
    security_rows = extract_spec_rows(workbook, "63 Security Cabins")
    security_electrical = next(
        row["detail"] for row in security_rows if row["component"] == "Electrical fittings"
    )

    set_detail(specs["low-cost-porta-cabin"], "Wall insulation", "25 mm glass wool", "L16 amendment §4.2")
    set_detail(specs["low-cost-porta-cabin"], "Roof insulation", "50 mm glass wool", "L16 amendment §4.2")
    set_detail(specs["mini-porta-cabin"], "Wall insulation", "12 mm heatlon", "L16 amendment §4.2")
    set_detail(specs["mini-porta-cabin"], "Roof insulation", "25 mm glass wool", "L16 amendment §4.2")
    set_detail(
        specs["mini-porta-cabin"],
        "Electrical fittings",
        security_electrical,
        "63 Security Cabins!Electrical fittings",
    )

    ms_values = {
        "Interior walls": "8–10 mm fibre-cement board",
        "Ceiling": "8 mm fibre-cement ceiling",
        "Floor base": "24 mm cement board",
        "Floor finish": "2–3 mm commercial PVC or epoxy",
        "Main door / service door": "Heavy single-leaf MS door with industrial lockset",
    }
    steel_values = {
        "Interior walls": "0.50 mm pre-painted metal liner",
        "Ceiling": "0.50 mm metal liner",
        "Floor base": "heavy MS floor plate",
        "Floor finish": "3 mm chequered plate",
        "Main door / service door": "Heavy double-leaf MS door with industrial lockset",
    }
    for component, detail in ms_values.items():
        set_detail(specs["ms-porta-cabin"], component, detail, "L16 amendment §4.1")
    for component, detail in steel_values.items():
        set_detail(specs["steel-porta-cabin"], component, detail, "L16 amendment §4.1")


def validate_specs(specs: dict[str, list[dict[str, Any]]]) -> dict[str, dict[str, Any]]:
    as_maps = {
        slug: {row["component"]: row["detail"] for row in rows}
        for slug, rows in specs.items()
    }
    hub = as_maps["porta-cabins"]
    proof: dict[str, dict[str, Any]] = {}
    for slug, (expected_hub, sibling, expected_sibling) in EXPECTED_DIVERGENCE.items():
        hub_diff = sum(as_maps[slug][name] != hub[name] for name in hub)
        sibling_diff = sum(
            as_maps[slug][name] != as_maps[sibling][name] for name in as_maps[slug]
        )
        if (hub_diff, sibling_diff) != (expected_hub, expected_sibling):
            raise SystemExit(
                f"Divergence mismatch for {slug}: hub={hub_diff}, sibling={sibling_diff}; "
                f"expected hub={expected_hub}, sibling={expected_sibling}"
            )
        proof[slug] = {
            "hubRows": hub_diff,
            "hubPercent": round(hub_diff / 30 * 100, 1),
            "nearestSibling": sibling,
            "siblingRows": sibling_diff,
        }

    for component in HARD_COMMON:
        values = {as_maps[slug][component] for slug in as_maps}
        if len(values) != 1:
            raise SystemExit(f"Hard-common row drift: {component}")

    for slug, rows in specs.items():
        for row in rows:
            row["differsFromHub"] = row["detail"] != hub[row["component"]]
    return proof


def extract_prices(workbook: openpyxl.Workbook) -> dict[str, list[dict[str, Any]]]:
    sheet = workbook["576 Pricing Matrix"]
    by_name = {meta["name"]: slug for slug, meta in PRODUCTS.items()}
    found: dict[str, dict[str, dict[str, Any]]] = {slug: {} for slug in PRODUCTS}
    for row in sheet.iter_rows(min_row=5, values_only=True):
        product_name = row[3]
        if product_name not in by_name:
            continue
        published_size = str(row[5])
        size_slug = published_size.split("x8.5 ft", 1)[0]
        size_slug = "x".join(published_size.split("x")[:2])
        slug = by_name[product_name]
        found[slug][size_slug] = {
            "sizeSlug": size_slug,
            "variantSku": str(row[0]),
            "publishedSize": published_size,
            "areaSqft": int(row[6]),
            "ratePerSqft": round_rupee(row[8]),
            "priceExGst": round_rupee(row[9]),
            "priceInclGst": round_rupee(row[16]),
            "costingStatus": str(row[17]),
        }
    ordered: dict[str, list[dict[str, Any]]] = {}
    for slug in PRODUCTS:
        order = MINI_SIZE_ORDER if slug == "mini-porta-cabin" else SIZE_ORDER
        missing = [size for size in order if size not in found[slug]]
        if missing:
            raise SystemExit(f"Missing price rows for {slug}: {missing}")
        ordered[slug] = [found[slug][size] for size in order]

    expected_20x10 = {
        "porta-cabins": 250000,
        "low-cost-porta-cabin": 240000,
        "luxury-porta-cabin": 370000,
        "mini-porta-cabin": 240000,
        "ms-porta-cabin": 360000,
        "porta-cabin-shop": 280000,
        "porta-cabin-with-toilet": 300000,
        "portacabin-office": 290000,
        "steel-porta-cabin": 360000,
    }
    for slug, expected in expected_20x10.items():
        actual = next(row["priceExGst"] for row in ordered[slug] if row["sizeSlug"] == "20x10")
        if actual != expected:
            raise SystemExit(f"20x10 draft cross-check failed for {slug}: {actual} != {expected}")
    return ordered


def update_variant_files(prices: dict[str, list[dict[str, Any]]]) -> None:
    data_dir = ROOT / "src" / "data" / "products"
    for slug in PRODUCTS:
        target = data_dir / f"{slug}.json"
        source = target
        if slug == "portacabin-office" and not source.exists():
            source = data_dir / "porta-cabin-office.json"
        data = json.loads(source.read_text(encoding="utf-8"))
        data["productSlug"] = slug
        if slug == "portacabin-office":
            data["productName"] = "Portacabin Office"
            data["applicationsDataset"] = "portacabin-office"
        data["specPdfHref"] = f"/specs/{slug}-technical-specification.pdf"
        data.pop("pricePerSqft", None)
        data.pop("emitAggregateOffer", None)

        price_by_size = {row["sizeSlug"]: row for row in prices[slug]}
        variants = [
            variant
            for variant in data["variants"]
            if variant["sizeSlug"] in price_by_size
        ]
        variants.sort(key=lambda variant: list(price_by_size).index(variant["sizeSlug"]))
        for variant in variants:
            price = price_by_size[variant["sizeSlug"]]
            variant["priceExGst"] = price["priceExGst"]
            variant["priceInclGst"] = price["priceInclGst"]
        data["variants"] = variants
        if data["defaultVariant"] not in price_by_size:
            data["defaultVariant"] = "20x10"
        target.write_text(
            json.dumps(data, ensure_ascii=False, indent=2) + "\n",
            encoding="utf-8",
            newline="\n",
        )


def parse_section_h(draft_path: Path) -> dict[str, Any]:
    text = draft_path.read_text(encoding="utf-8")
    start = text.index("## 8 · §H COPY")
    end = text.index("## 10 · SUPPORTING-KEYWORD SECTION")
    section = text[start:end]
    page_pattern = re.compile(
        r"^### /product/porta-cabins/(?P<slug>[a-z0-9-]+).*?\n"
        r"(?P<body>.*?)(?=^### /product/porta-cabins/|\Z)",
        re.MULTILINE | re.DOTALL,
    )
    result: dict[str, Any] = {}
    for page_match in page_pattern.finditer(section):
        slug = page_match.group("slug")
        body = page_match.group("body")
        h2_match = re.search(r"^\*\*H2 \(\d+c\):\*\* `([^`]+)`", body, re.MULTILINE)
        guide_match = re.search(
            r"^\*\*Guidance line \(\d+c\):\*\* `([^`]+)`", body, re.MULTILINE
        )
        if not h2_match or not guide_match:
            raise SystemExit(f"Missing Section H heading/guidance for {slug}")
        dataset: dict[str, Any] = {
            "h2": h2_match.group(1),
            "guidanceLine": guide_match.group(1),
        }
        tab_pattern = re.compile(
            r"^#### Tab \d+ — (?P<size>\d+×\d+) ft\n"
            r"\*\*Title \(\d+c\):\*\* `(?P<title>[^`]+)`\n\n"
            r"\*\*Body \(\d+c\):\*\*\n> (?P<paragraph>.+?)\n\n"
            r"\*\*Uses:\*\* (?P<uses>.+?)\n\n"
            r"\*\*Stats:\*\* `(?P<stats>[^`]+)`",
            re.MULTILINE | re.DOTALL,
        )
        for tab in tab_pattern.finditer(body):
            size_slug = tab.group("size").replace("×", "x")
            uses = [
                item.strip()
                for item in re.split(r"\s*✓\s*", tab.group("uses"))
                if item.strip()
            ]
            if len(uses) != 4:
                raise SystemExit(f"{slug}/{size_slug}: expected four uses, found {uses}")
            dataset[size_slug] = {
                "h2": tab.group("title"),
                "intro": tab.group("paragraph").strip(),
                "h3": "Uses",
                "applications": uses,
                "stats": tab.group("stats"),
            }
        expected = 4 if slug == "mini-porta-cabin" else 9
        actual = len([key for key in dataset if re.fullmatch(r"\d+x\d+", key)])
        if actual != expected:
            raise SystemExit(f"{slug}: expected {expected} Section H tabs, found {actual}")
        result[slug] = dataset
    expected_slugs = {
        "mini-porta-cabin",
        "ms-porta-cabin",
        "steel-porta-cabin",
        "porta-cabin-shop",
        "porta-cabin-with-toilet",
        "portacabin-office",
    }
    if set(result) != expected_slugs:
        raise SystemExit(f"Section H slug mismatch: {set(result)}")
    return result


def update_section_h(draft_path: Path) -> None:
    path = ROOT / "src" / "data" / "products" / "section-h-datasets.json"
    current = json.loads(path.read_text(encoding="utf-8"))
    current.update(parse_section_h(draft_path))
    path.write_text(
        json.dumps(current, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
        newline="\n",
    )


def register_pdf_fonts() -> tuple[str, str]:
    regular = Path("C:/Windows/Fonts/arial.ttf")
    bold = Path("C:/Windows/Fonts/arialbd.ttf")
    if regular.exists() and bold.exists():
        pdfmetrics.registerFont(TTFont("SamanArial", str(regular)))
        pdfmetrics.registerFont(TTFont("SamanArialBold", str(bold)))
        return "SamanArial", "SamanArialBold"
    return "Helvetica", "Helvetica-Bold"


def generate_pdf(
    slug: str,
    meta: dict[str, str],
    rows: list[dict[str, Any]],
    prices: list[dict[str, Any]],
    output: Path,
) -> None:
    regular_font, bold_font = register_pdf_fonts()
    forest = colors.HexColor("#173F31")
    leaf = colors.HexColor("#3E8E54")
    pale = colors.HexColor("#EEF6F1")
    line = colors.HexColor("#D5E2DB")
    ink = colors.HexColor("#26332D")
    muted = colors.HexColor("#5F6F67")

    styles = getSampleStyleSheet()
    normal = ParagraphStyle(
        "Body",
        parent=styles["BodyText"],
        fontName=regular_font,
        fontSize=8.2,
        leading=10.5,
        textColor=ink,
        alignment=TA_LEFT,
    )
    small = ParagraphStyle(
        "Small",
        parent=normal,
        fontSize=7.2,
        leading=9,
        textColor=muted,
    )
    h1 = ParagraphStyle(
        "H1",
        parent=styles["Heading1"],
        fontName=bold_font,
        fontSize=17,
        leading=20,
        textColor=forest,
        spaceAfter=3 * mm,
    )
    h2 = ParagraphStyle(
        "H2",
        parent=styles["Heading2"],
        fontName=bold_font,
        fontSize=12,
        leading=14,
        textColor=forest,
        spaceBefore=3 * mm,
        spaceAfter=2 * mm,
    )
    cell = ParagraphStyle("Cell", parent=normal, fontSize=7, leading=8.5)
    cell_bold = ParagraphStyle(
        "CellBold", parent=cell, fontName=bold_font, textColor=forest
    )
    header_cell = ParagraphStyle(
        "HeaderCell",
        parent=cell,
        fontName=bold_font,
        textColor=colors.white,
    )

    def header_footer(canvas: Any, doc: Any) -> None:
        canvas.saveState()
        width, height = A4
        canvas.setFillColor(forest)
        canvas.rect(0, height - 25 * mm, width, 25 * mm, stroke=0, fill=1)
        canvas.setFillColor(colors.white)
        canvas.setFont(bold_font, 15)
        canvas.drawString(18 * mm, height - 12 * mm, "SAMAN PORTABLE")
        canvas.setFont(regular_font, 7.5)
        canvas.setFillColor(colors.HexColor("#D5E8DE"))
        canvas.drawString(
            18 * mm,
            height - 18 * mm,
            "Factory-built porta cabins & prefab structures  ·  Bengaluru & Greater Noida  ·  Pan-India delivery",
        )
        canvas.setFont(bold_font, 8)
        canvas.setFillColor(colors.white)
        canvas.drawRightString(width - 18 * mm, height - 12 * mm, meta["name"].upper())
        canvas.setFont(regular_font, 7)
        canvas.drawRightString(
            width - 18 * mm, height - 18 * mm, "Technical Specification & Price Sheet"
        )
        canvas.setStrokeColor(leaf)
        canvas.setLineWidth(1.2 * mm)
        canvas.line(0, height - 25 * mm, width, height - 25 * mm)
        canvas.setStrokeColor(line)
        canvas.setLineWidth(0.3)
        canvas.line(18 * mm, 14 * mm, width - 18 * mm, 14 * mm)
        canvas.setFillColor(muted)
        canvas.setFont(regular_font, 6.5)
        canvas.drawString(18 * mm, 9 * mm, "SAMAN POS India Private Limited")
        canvas.drawCentredString(width / 2, 9 * mm, "www.samanportable.com")
        canvas.drawRightString(width - 18 * mm, 9 * mm, f"Page {doc.page}")
        canvas.restoreState()

    doc = BaseDocTemplate(
        str(output),
        pagesize=A4,
        leftMargin=18 * mm,
        rightMargin=18 * mm,
        topMargin=32 * mm,
        bottomMargin=19 * mm,
        title=f"{meta['name']} Technical Specification",
        author="SAMAN POS India Private Limited",
    )
    frame = Frame(
        doc.leftMargin,
        doc.bottomMargin,
        doc.width,
        doc.height,
        id="content",
        leftPadding=0,
        rightPadding=0,
        topPadding=0,
        bottomPadding=0,
    )
    doc.addPageTemplates([PageTemplate(id="saman", frames=[frame], onPage=header_footer)])

    story: list[Any] = [
        Paragraph(f"{meta['name']} — Technical Specifications & Price List", h1),
        Paragraph(f"<b>Canonical URL:</b> {meta['canonical']}", small),
        Spacer(1, 2 * mm),
        Paragraph(
            "Newly fabricated at our own works from MS sheet, MS pipe framing and aluminium sections — not a converted shipping container.",
            normal,
        ),
        Spacer(1, 2 * mm),
        Paragraph("Sizes & Prices", h2),
    ]

    price_data = [
        [
            Paragraph("SIZE", header_cell),
            Paragraph("AREA", header_cell),
            Paragraph("RATE / SQ FT", header_cell),
            Paragraph("EX-GST", header_cell),
            Paragraph("INCL. 18% GST", header_cell),
        ]
    ]
    for price in prices:
        price_data.append(
            [
                Paragraph(price["publishedSize"].replace("x8.5 ft", "×8.5 ft"), cell_bold),
                Paragraph(f"{price['areaSqft']} sq ft", cell),
                Paragraph(f"₹{indian(price['ratePerSqft'])}", cell),
                Paragraph(f"₹{indian(price['priceExGst'])}", cell_bold),
                Paragraph(f"₹{indian(price['priceInclGst'])}", cell),
            ]
        )
    price_table = Table(
        price_data,
        colWidths=[46 * mm, 25 * mm, 32 * mm, 34 * mm, 37 * mm],
        repeatRows=1,
    )
    price_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), forest),
                ("GRID", (0, 0), (-1, -1), 0.3, line),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, pale]),
            ]
        )
    )
    story.extend(
        [
            price_table,
            Spacer(1, 1.5 * mm),
            Paragraph(
                "Base specification price — customisations quoted separately. Prices are supply-only and ex-factory; freight and installation are confirmed at quotation.",
                small,
            ),
            Paragraph("30-row Technical Specification", h2),
            Paragraph(
                "Rows marked “DIFFERS” are different from the Porta Cabins reference specification.",
                small,
            ),
            Spacer(1, 1.5 * mm),
        ]
    )

    current_group = None
    for row in rows:
        if row["group"] != current_group:
            current_group = row["group"]
            story.append(Paragraph(current_group, h2))
        marker = "DIFFERS" if row["differsFromHub"] else ""
        table = Table(
            [
                [
                    Paragraph(row["component"], cell_bold),
                    Paragraph(row["detail"], cell),
                    Paragraph(marker, cell_bold),
                ]
            ],
            colWidths=[42 * mm, 113 * mm, 19 * mm],
        )
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, -1), pale if marker else colors.white),
                    ("BOX", (0, 0), (-1, -1), 0.3, line),
                    ("INNERGRID", (0, 0), (-1, -1), 0.3, line),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 4),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                    ("TOPPADDING", (0, 0), (-1, -1), 4),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ]
            )
        )
        story.append(table)

    facts = [
        Paragraph("Commercial & Company Facts", h2),
        Paragraph(f"<b>Warranty:</b> {WARRANTY_SENTENCE}", normal),
        Paragraph(
            "<b>Turnaround:</b> Delivery in 7–21 working days; fixed-price quotation in 48 hours.",
            normal,
        ),
        Paragraph(
            "<b>Certifications:</b> ISO 9001:2015 · ISO 14001:2015 · ISO 45001:2018 · "
            "NSIC SPRS · Udyam UDYAM-KR-03-0172770 · ZED Bronze · DPIIT Startup India · GST registered.",
            normal,
        ),
        Paragraph(
            "<b>Bengaluru (South):</b> +91 88616 22859 · +91 80886 85440 · sales@samanportable.com",
            normal,
        ),
        Paragraph(
            "<b>Greater Noida (North):</b> +91 87960 39938 · +91 97089 89937 · ncr@samanportable.com",
            normal,
        ),
        Paragraph("<b>Generated:</b> 26 July 2026", normal),
    ]
    story.append(KeepTogether(facts))
    doc.build(story)
    if output.stat().st_size > 400_000:
        raise SystemExit(f"{output.name} exceeds 400 KB: {output.stat().st_size}")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", required=True, type=Path)
    parser.add_argument("--draft", required=True, type=Path)
    args = parser.parse_args()
    verify_workbook(args.workbook)

    workbook = openpyxl.load_workbook(args.workbook, read_only=True, data_only=True)
    specs = {
        slug: extract_spec_rows(workbook, meta["sheet"])
        for slug, meta in PRODUCTS.items()
    }
    apply_amendment(workbook, specs)
    proof = validate_specs(specs)
    prices = extract_prices(workbook)

    generated = {
        "sourceWorkbook": args.workbook.name,
        "sourceWorkbookMd5": EXPECTED_MD5,
        "rowCount": 30,
        "hardCommonRows": sorted(HARD_COMMON),
        "divergenceProof": proof,
        "products": {
            slug: {
                **PRODUCTS[slug],
                "specifications": specs[slug],
                "prices": prices[slug],
            }
            for slug in PRODUCTS
        },
    }
    generated_path = ROOT / "src" / "data" / "products" / "c01-specifications.json"
    generated_path.write_text(
        json.dumps(generated, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
        newline="\n",
    )
    update_variant_files(prices)
    update_section_h(args.draft)

    pdf_dir = ROOT / "public" / "specs"
    pdf_dir.mkdir(parents=True, exist_ok=True)
    for slug, meta in PRODUCTS.items():
        generate_pdf(
            slug,
            meta,
            specs[slug],
            prices[slug],
            pdf_dir / f"{slug}-technical-specification.pdf",
        )

    print(f"Generated {generated_path.relative_to(ROOT)}")
    for slug in PRODUCTS:
        if slug != "porta-cabins":
            item = proof[slug]
            print(
                f"{slug}: {item['hubRows']}/30 = {item['hubPercent']}%; "
                f"nearest {item['nearestSibling']} = {item['siblingRows']}"
            )
    for pdf in sorted(pdf_dir.glob("*-technical-specification.pdf")):
        print(f"{pdf.relative_to(ROOT)} {pdf.stat().st_size} bytes")


if __name__ == "__main__":
    main()

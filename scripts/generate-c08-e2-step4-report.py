#!/usr/bin/env python3
"""Generate the source-backed C08-E2 pricing, specification, and heading report."""

from __future__ import annotations

import hashlib
import json
import sys
from pathlib import Path
from urllib.request import Request, urlopen

from openpyxl import load_workbook


REPO = Path(__file__).resolve().parents[1]
SPEC_PATH = REPO / "src" / "data" / "products" / "c08-specifications.json"
REPORT_PATH = REPO / "page-structure" / "C08" / "C08-E2-STEP4-REPORT-05Aug2026.json"
WORKBOOK_PATH = Path(
    r"D:\Project-shekhar\all-technical-specifications\00-FINAL-FILES"
    r"\SAMAN_MASTER_64_Products_Detailed_Technical_Specs_9_Sizes_Report-with-price- PR.xlsx"
)
ROUTES = {
    "container-houses": "/product/container-houses",
    "prefab-container-homes": "/product/container-houses/prefab-container-homes",
    "luxury-container-houses": "/product/container-houses/luxury-container-houses",
    "shipping-container-homes": "/product/container-houses/shipping-container-homes",
    "affordable-container-homes": "/product/container-houses/affordable-container-homes",
    "prefabricated-container-house": "/product/container-houses/prefabricated-container-house",
}
HARD_COMMON = (
    "Electrical protection",
    "Electrical wiring",
    "Fasteners & sealing",
    "Grills / mosquito mesh",
    "Warranty",
    "Welding & fabrication",
)


class HeadingParser(__import__("html.parser").parser.HTMLParser):
    def __init__(self) -> None:
        super().__init__(convert_charrefs=True)
        self.current: tuple[str, list[str]] | None = None
        self.headings: list[dict[str, str]] = []

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        if tag in {"h1", "h2", "h3", "h4", "h5", "h6"}:
            self.current = (tag, [])

    def handle_endtag(self, tag: str) -> None:
        if self.current and tag == self.current[0]:
            import re

            text = re.sub(r"\s+", " ", "".join(self.current[1])).strip()
            self.headings.append({"level": tag.upper(), "text": text})
            self.current = None

    def handle_data(self, data: str) -> None:
        if self.current:
            self.current[1].append(data)


def md5(path: Path) -> str:
    digest = hashlib.md5()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def rendered_headings(base_url: str, route: str) -> tuple[int, list[dict[str, str]]]:
    request = Request(base_url.rstrip("/") + route, headers={"User-Agent": "C08-E2-report/1.0"})
    with urlopen(request, timeout=30) as response:
        parser = HeadingParser()
        parser.feed(response.read().decode("utf-8", errors="replace"))
        return response.status, parser.headings


def pricing_evidence() -> dict[str, object]:
    workbook = load_workbook(WORKBOOK_PATH, read_only=True, data_only=True)
    sheet = workbook["576 Pricing Matrix"]
    exact_rows = []
    for row_number, row in enumerate(sheet.iter_rows(min_row=5, values_only=True), start=5):
        if str(row[3] or "").strip().casefold() == "prefabricated container house":
            exact_rows.append(
                {
                    "row": row_number,
                    "productName": row[3],
                    "publishedSize": row[5],
                    "baseRatePerSqFt": row[8],
                    "baseCabinPrice": row[9],
                    "costingStatus": row[17],
                }
            )
    return {
        "sourceWorkbook": str(WORKBOOK_PATH),
        "sourceBytes": WORKBOOK_PATH.stat().st_size,
        "sourceMd5": md5(WORKBOOK_PATH),
        "sheet": "576 Pricing Matrix",
        "exactProductName": "Prefabricated Container House",
        "exactRowCount": len(exact_rows),
        "rows": exact_rows,
        "conclusion": (
            "No priced rows exist for prefabricated-container-house; no sibling prices may be mapped."
            if not exact_rows
            else "Exact priced rows exist."
        ),
    }


def main() -> None:
    base_url = sys.argv[1] if len(sys.argv) > 1 else "http://127.0.0.1:3042"
    dataset = json.loads(SPEC_PATH.read_text(encoding="utf-8"))["products"]
    component_tables = []
    heading_reports = []
    for slug, route in ROUTES.items():
        entry = dataset.get(slug)
        rows = (
            [
                {"component": row["component"], "detail": row["detail"]}
                for row in entry["specifications"]
            ]
            if entry
            else []
        )
        component_tables.append(
            {
                "route": route,
                "source": SPEC_PATH.relative_to(REPO).as_posix() if entry else None,
                "rowCount": len(rows),
                "rows": rows,
                "note": None if entry else "No C-08 30-row specification table exists for this route.",
            }
        )
        status, headings = rendered_headings(base_url, route)
        heading_reports.append(
            {
                "route": route,
                "status": status,
                "h2": [heading["text"] for heading in headings if heading["level"] == "H2"],
                "fullHeadingStructure": headings,
            }
        )

    common_proof = []
    for component in HARD_COMMON:
        values = {}
        for slug in ROUTES:
            if slug in dataset:
                values[slug] = next(
                    row["detail"]
                    for row in dataset[slug]["specifications"]
                    if row["component"] == component
                )
        common_proof.append(
            {
                "component": component,
                "detail": next(iter(values.values())),
                "fiveGeneratedRoutesByteIdentical": len(set(values.values())) == 1,
                "prefabricatedContainerHouseHasRow": False,
            }
        )

    report = {
        "generatedFromPreview": base_url,
        "pricingMatrix": pricing_evidence(),
        "componentTables": component_tables,
        "hardCommonRows": {
            "expected": list(HARD_COMMON),
            "proof": common_proof,
            "allFiveGeneratedRoutesByteIdentical": all(
                row["fiveGeneratedRoutesByteIdentical"] for row in common_proof
            ),
            "allSixRoutesByteIdentical": False,
            "reason": "prefabricated-container-house has no C-08 30-row specification table, so six-route identity cannot be asserted.",
        },
        "headings": heading_reports,
    }
    REPORT_PATH.parent.mkdir(parents=True, exist_ok=True)
    REPORT_PATH.write_text(
        json.dumps(report, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
        newline="\n",
    )
    print(REPORT_PATH.relative_to(REPO).as_posix())
    print(
        f"pricing rows={report['pricingMatrix']['exactRowCount']}; "
        f"spec rows={[row['rowCount'] for row in component_tables]}; "
        f"route 200={sum(row['status'] == 200 for row in heading_reports)}/6"
    )


if __name__ == "__main__":
    main()

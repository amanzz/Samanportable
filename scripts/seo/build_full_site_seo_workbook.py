#!/usr/bin/env python3
"""Build the implementation workbook for the SAMAN Google-policy URL audit."""

from __future__ import annotations

import csv
import hashlib
import json
import math
import re
import subprocess
from collections import Counter, defaultdict
from datetime import date
from pathlib import Path
from typing import Any, Iterable
from urllib.parse import urlsplit

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.comments import Comment
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

from full_site_policy_crawl import ORIGIN, load_redirects, normalize_url


AUDIT_DATE = date(2026, 8, 11)
GOOGLE_HELPFUL = "https://developers.google.com/search/docs/fundamentals/creating-helpful-content"
GOOGLE_CORE = "https://developers.google.com/search/docs/appearance/core-updates"
GOOGLE_SPAM = "https://developers.google.com/search/docs/essentials/spam-policies"
GOOGLE_CANONICAL = "https://developers.google.com/search/docs/crawling-indexing/consolidate-duplicate-urls"
GOOGLE_PAGE_EXPERIENCE = "https://developers.google.com/search/docs/appearance/page-experience"
GOOGLE_RANKING = "https://developers.google.com/search/docs/appearance/ranking-systems-guide"
GOOGLE_UPDATES = "https://status.search.google.com/products/rGHU1u87FJnkP6W2GwMi/history"
ALLOWED_PHONES = {"918861622859", "918796039938"}
BANNED_PHONE = "916200909435"

GREEN = "1F4E3D"
GREEN_2 = "DDEDE5"
GREEN_3 = "EAF4EF"
GOLD = "D9A441"
RED = "C00000"
RED_FILL = "FCE4D6"
AMBER_FILL = "FFF2CC"
BLUE_FILL = "DDEBF7"
GREY_FILL = "E7E6E6"
WHITE = "FFFFFF"
BLACK = "000000"
THIN_GREY = Side(style="thin", color="D9E1E8")

CLUSTER_LABELS = {
    "container-cafe": "Container Cafe",
    "container-houses": "Container Houses",
    "container-offices": "Container Offices",
    "industrial-sheds": "Industrial Sheds",
    "labor-colony": "Labor Colony",
    "peb-constructions": "PEB Constructions",
    "porta-cabins": "Porta Cabins",
    "portable-cabin": "Portable Cabin",
    "portable-office": "Portable Office",
    "portable-toilet": "Portable Toilet",
    "pre-engineered-buildings": "Pre-Engineered Buildings",
    "prefab-buildings": "Prefab Buildings",
    "prefabricated-houses": "Prefabricated Houses",
    "security-cabins": "Security Cabins",
    "sandwich-panel": "Sandwich Panels",
    "roofing-sheet": "Roofing Sheets",
    "wall-sheet": "Wall Sheets",
}

HUB_TARGETS = {
    "Container Cafe": f"{ORIGIN}/product/container-cafe",
    "Container Houses": f"{ORIGIN}/product/container-houses",
    "Container Offices": f"{ORIGIN}/product/container-offices",
    "Industrial Sheds": f"{ORIGIN}/product/industrial-sheds",
    "Labor Colony": f"{ORIGIN}/product/labor-colony",
    "PEB Constructions": f"{ORIGIN}/product/peb-constructions",
    "Porta Cabins": f"{ORIGIN}/product/porta-cabins",
    "Portable Cabin": f"{ORIGIN}/product/portable-cabin",
    "Portable Office": f"{ORIGIN}/product/portable-office",
    "Portable Toilet": f"{ORIGIN}/product/portable-toilet",
    "Pre-Engineered Buildings": f"{ORIGIN}/product/pre-engineered-buildings",
    "Prefab Buildings": f"{ORIGIN}/product/prefab-buildings",
    "Prefabricated Houses": f"{ORIGIN}/product/prefabricated-houses",
    "Security Cabins": f"{ORIGIN}/product/security-cabins",
    "Sandwich Panels": f"{ORIGIN}/product/sandwich-panel",
    "Roofing Sheets": f"{ORIGIN}/product/roofing-sheet",
    "Wall Sheets": f"{ORIGIN}/product/wall-sheet",
}

ANGLES = {
    "Container Cafe": "food-and-beverage business layout, service window, utilities and customer-facing fit-out",
    "Container Houses": "container-based residential planning, insulation, rooms and long-stay comfort",
    "Container Offices": "heavy-duty container-based project office, lifting, relocation and industrial use",
    "Industrial Sheds": "factory, warehouse or workshop span, roofing, cladding, loading and ventilation",
    "Labor Colony": "project-scale worker accommodation, rooms, hygiene, kitchens, toilets and circulation",
    "PEB Constructions": "turnkey design, fabrication, supply, erection and project execution service",
    "Porta Cabins": "site-ready prefabricated cabin category for contractors and industrial projects",
    "Portable Cabin": "relocatable multi-use cabin, mobility, custom size and flexible installation",
    "Portable Office": "office-only workspace planning, staff comfort, electricals, AC and partitions",
    "Portable Toilet": "standalone sanitation, plumbing, drainage, cleaning access and hygiene",
    "Pre-Engineered Buildings": "engineered steel building system, clear spans, frames, purlins and cladding",
    "Prefab Buildings": "factory-made commercial and institutional modular building systems",
    "Prefabricated Houses": "residential prefab homes, family layout, insulation, foundation and finishes",
    "Security Cabins": "compact guard booth visibility, windows, counter, ventilation and gate placement",
    "Sandwich Panels": "panel construction, core option, facing, joint, application and quotation-confirmed performance",
    "Roofing Sheets": "roof sheet material, profile, thickness, fasteners, drainage and application",
    "Wall Sheets": "wall sheet material, profile, finish, fixing, joint and application",
}

SPEC_TOPICS = {
    "Container Cafe": "base/frame; wall and roof; kitchen/service layout; electrical load; plumbing/drainage; ventilation; floor; openings; transport",
    "Container Houses": "base/frame; wall/roof insulation; room layout; doors/windows; electrical; plumbing; floor/finish; foundation; transport/installation",
    "Container Offices": "container shell/base; frame; wall/roof insulation; floor; doors/windows; electrical/AC; lifting points; transport; stacking limits if verified",
    "Industrial Sheds": "design basis; span/height; columns/trusses; purlins/bracing; roofing/cladding; loads; ventilation; foundation interface; erection scope",
    "Labor Colony": "module/frame; room capacity; wall/roof; floor; doors/windows; electrical; toilets/plumbing; kitchen; drainage; fire/safety only if verified",
    "PEB Constructions": "design basis; primary/secondary steel; loads; roofing/cladding; bracing; foundation interface; fabrication QA; erection; exclusions",
    "Porta Cabins": "MS frame; wall panel; roof; floor; door/window; electrical/AC; plumbing where applicable; dimensions; lifting/transport; installation base",
    "Portable Cabin": "frame; wall/roof; floor; openings; electrical; size/layout; relocation method; transport; installation base; customization",
    "Portable Office": "frame; insulation; floor/ceiling; workstations/partitions; doors/windows; lighting/electrical; AC provision; dimensions; delivery/setup",
    "Portable Toilet": "structure; wall/roof; floor; sanitary fixtures; water inlet/storage; waste outlet/tank; ventilation; cleaning access; lifting/placement",
    "Pre-Engineered Buildings": "design basis; clear span; primary frames; purlins/girts; bracing; roof/wall cladding; loads; coatings; erection interface",
    "Prefab Buildings": "structural system; wall/roof; floor; modules; openings; electrical/plumbing; insulation; foundation; transport/installation",
    "Prefabricated Houses": "structural system; wall/roof insulation; residential layout; floor/finish; doors/windows; electrical; plumbing; foundation; installation",
    "Security Cabins": "frame; wall/roof; floor; door/windows/visibility; counter; electrical/fan; dimensions; lifting/placement; weather protection",
    "Sandwich Panels": "core type; facing; thickness; width/length; joint profile; fasteners; weight; application; performance values only from approved data",
    "Roofing Sheets": "material; profile; thickness; coating/finish; effective cover; length; fasteners; slope/drainage; installation and maintenance",
    "Wall Sheets": "material; profile; thickness; coating/finish; effective cover; length; joint/fasteners; application; installation and maintenance",
}


def compact(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def clean_phone(value: str) -> str:
    digits = re.sub(r"\D", "", value)
    if len(digits) == 10:
        digits = "91" + digits
    return digits


def slug_words(url: str) -> str:
    slug = urlsplit(url).path.rstrip("/").split("/")[-1] or "home"
    return re.sub(r"[-_]+", " ", slug).strip().title()


def classify_page(url: str, record: dict[str, Any]) -> tuple[str, str]:
    path = urlsplit(url).path.rstrip("/") or "/"
    sources = set(record.get("sources", []))
    if path == "/":
        return "Homepage", "Brand / Site Router"
    if "sitemap-locations" in sources:
        return "Local SEO Page", infer_cluster(url)
    if "sitemap-products" in sources or path.startswith("/product/"):
        segments = [p for p in path.split("/") if p]
        return ("Product Hub" if len(segments) == 2 else "Product Subpage"), infer_cluster(url)
    if "sitemap-projects" in sources or path.startswith("/projects"):
        return "Project / Case Study", infer_cluster(url)
    if "sitemap-editorial" in sources:
        return "Editorial / Guide", infer_cluster(url)
    if path in {"/about", "/contact", "/privacy-policy", "/terms-and-conditions", "/shipping-policy", "/return-refund-policy", "/blog", "/product"}:
        return "Static / Trust / Hub", infer_cluster(url)
    return "Legacy / Other", infer_cluster(url)


def infer_cluster(url: str) -> str:
    path = urlsplit(url).path.lower()
    ordered = [
        ("portable-office", "Portable Office"), ("container-office", "Container Offices"),
        ("porta-cabin", "Porta Cabins"), ("portacabin", "Porta Cabins"),
        ("portable-cabin", "Portable Cabin"), ("container-house", "Container Houses"),
        ("container-home", "Container Houses"), ("container-cafe", "Container Cafe"),
        ("container-coffee", "Container Cafe"), ("container-restaurant", "Container Cafe"),
        ("portable-toilet", "Portable Toilet"), ("toilet-cabin", "Portable Toilet"),
        ("security-cabin", "Security Cabins"), ("guard-cabin", "Security Cabins"),
        ("labor-colony", "Labor Colony"), ("labour-colony", "Labor Colony"),
        ("labor-camp", "Labor Colony"), ("labour-camp", "Labor Colony"),
        ("labor-hut", "Labor Colony"), ("labour-hut", "Labor Colony"),
        ("industrial-shed", "Industrial Sheds"), ("warehouse-shed", "Industrial Sheds"),
        ("peb-construction", "PEB Constructions"),
        ("pre-engineered-building", "Pre-Engineered Buildings"),
        ("prefabricated-house", "Prefabricated Houses"), ("prefab-house", "Prefabricated Houses"),
        ("prefab-building", "Prefab Buildings"), ("prefabricated-building", "Prefab Buildings"),
        ("sandwich-panel", "Sandwich Panels"), ("puf-panel", "Sandwich Panels"),
        ("pir-panel", "Sandwich Panels"), ("eps-panel", "Sandwich Panels"),
        ("rockwool-panel", "Sandwich Panels"), ("glass-wool-panel", "Sandwich Panels"),
        ("roofing-sheet", "Roofing Sheets"), ("roof-sheet", "Roofing Sheets"),
        ("wall-sheet", "Wall Sheets"),
    ]
    for needle, cluster in ordered:
        if needle in path:
            return cluster
    return "Other / Brand"


def load_gsc(root: Path) -> dict[str, dict[str, dict[str, float]]]:
    result: dict[str, dict[str, dict[str, float]]] = defaultdict(dict)
    paths = {
        "16M": root / "GSC_REPORTS/LAST_16_MONTHS/GSC_16M_Pages.csv",
        "Mid": root / "GSC_REPORTS/LAST_6_MONTHS/GSC_6M_Pages.csv",
        "Recent": root / "GSC_REPORTS/LAST_3_MONTHS/GSC_3M_Pages.csv",
    }
    for label, path in paths.items():
        accum: dict[str, dict[str, float]] = defaultdict(lambda: {"clicks": 0, "impressions": 0, "position_weighted": 0})
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            for row in csv.DictReader(handle):
                url = normalize_url(row.get("Top pages", ""))
                if not url:
                    continue
                clicks = float(row.get("Clicks", 0) or 0)
                impressions = float(row.get("Impressions", 0) or 0)
                position = float(row.get("Position", 0) or 0)
                accum[url]["clicks"] += clicks
                accum[url]["impressions"] += impressions
                accum[url]["position_weighted"] += position * max(impressions, 1)
        for url, values in accum.items():
            denom = max(values["impressions"], 1)
            result[url][label] = {
                "clicks": values["clicks"],
                "impressions": values["impressions"],
                "position": values["position_weighted"] / denom,
                "ctr": values["clicks"] / denom,
            }
    return result


def shingle_set(text: str, width: int = 8) -> set[int]:
    tokens = re.findall(r"[a-z0-9]+", text.lower())
    if len(tokens) < width:
        return {int(hashlib.blake2b(" ".join(tokens).encode(), digest_size=8).hexdigest(), 16)} if tokens else set()
    result = set()
    for index in range(len(tokens) - width + 1):
        value = " ".join(tokens[index:index + width]).encode("utf-8")
        result.add(int.from_bytes(hashlib.blake2b(value, digest_size=8).digest(), "big"))
    return result


def bottom_k(values: set[int], k: int = 128) -> set[int]:
    return set(sorted(values)[:k])


def similarity_map(records: dict[str, dict[str, Any]]) -> dict[str, dict[str, Any]]:
    candidates: list[tuple[str, str, str, set[int], set[int]]] = []
    for url, record in records.items():
        page_type, cluster = classify_page(url, record)
        if record.get("status") != 200 or not record.get("content_text"):
            continue
        if not any(source.startswith("sitemap-") and "images" not in source for source in record.get("sources", [])):
            continue
        shingles = shingle_set(record["content_text"])
        candidates.append((url, page_type, cluster, shingles, bottom_k(shingles)))
    by_group: dict[tuple[str, str], list[tuple[str, set[int], set[int]]]] = defaultdict(list)
    for url, page_type, cluster, shingles, sketch in candidates:
        broad_type = "Local" if page_type == "Local SEO Page" else ("Product" if page_type.startswith("Product") else page_type)
        by_group[(broad_type, cluster)].append((url, shingles, sketch))
    result: dict[str, dict[str, Any]] = {url: {"nearest": "", "similarity": 0.0} for url, *_ in candidates}
    for group in by_group.values():
        for index, (url_a, sh_a, sk_a) in enumerate(group):
            for url_b, sh_b, sk_b in group[index + 1:]:
                if not sk_a or not sk_b:
                    continue
                estimate = len(sk_a & sk_b) / max(1, min(len(sk_a), len(sk_b)))
                if estimate < 0.08:
                    continue
                union = len(sh_a | sh_b)
                exact = len(sh_a & sh_b) / union if union else 0
                if exact > result[url_a]["similarity"]:
                    result[url_a] = {"nearest": url_b, "similarity": exact}
                if exact > result[url_b]["similarity"]:
                    result[url_b] = {"nearest": url_a, "similarity": exact}
    return result


def redirect_chain(url: str, records: dict[str, dict[str, Any]], configured: dict[str, dict[str, Any]]) -> tuple[str, int, bool]:
    current = url
    seen: set[str] = set()
    hops = 0
    while hops < 12:
        if current in seen:
            return current, hops, True
        seen.add(current)
        target = ""
        row = configured.get(current)
        if row:
            target = row.get("destination_url") or ""
        if not target:
            location = records.get(current, {}).get("location", "")
            target = normalize_url(location) if location else ""
        if not target:
            return current, hops, False
        current = target
        hops += 1
    return current, hops, True


def primary_topic(url: str, record: dict[str, Any], page_type: str) -> str:
    h1 = compact(record.get("h1"))
    title = compact(record.get("title"))
    if h1 and len(h1) <= 100:
        value = re.split(r"\s+[—|]\s+", h1)[0]
    elif title:
        value = re.split(r"\s+[—|]\s+", title)[0]
    else:
        value = slug_words(url)
    value = re.sub(r"\b(best|top|no\.?\s*1|leading)\b", "", value, flags=re.I)
    value = compact(value).strip(" -–—|")
    return value[:95] or slug_words(url)


def location_from_topic(topic: str, url: str) -> str:
    match = re.search(r"\b(?:in|near|for)\s+([A-Z][A-Za-z .'-]+)$", topic)
    if match:
        return match.group(1).strip()
    slug = urlsplit(url).path.rstrip("/").split("/")[-1]
    match = re.search(r"-(?:in|near)-(.+)$", slug)
    return match.group(1).replace("-", " ").title() if match else "the location"


def proposed_title(topic: str, page_type: str, cluster: str, record: dict[str, Any]) -> str:
    current = compact(record.get("title"))
    if 28 <= len(current) <= 62 and not re.search(r"\b(best|top\s*\d+|no\.?\s*1)\b", current, re.I):
        return current
    has_price = bool(re.search(r"₹|\bprice\b|\bcost\b", record.get("content_text", ""), re.I))
    if page_type == "Homepage":
        candidate = "SAMAN Portable | Prefabricated Structures in India"
    elif page_type == "Local SEO Page":
        candidate = f"{topic} | Delivery & Specifications | SAMAN"
    elif page_type == "Product Hub":
        candidate = f"{topic} Manufacturer in India | SAMAN"
    elif page_type == "Product Subpage":
        middle = "Sizes, Specs & Price" if has_price else "Sizes & Specifications"
        candidate = f"{topic} | {middle} | SAMAN"
    elif page_type == "Editorial / Guide":
        candidate = f"{topic} | Practical Guide | SAMAN"
    else:
        candidate = current or f"{topic} | SAMAN Portable"
    if len(candidate) > 62:
        candidate = candidate.replace(" Manufacturer in India", " in India").replace("Specifications", "Specs")
    if len(candidate) > 62:
        candidate = candidate[:61].rstrip(" -|,")
    return candidate


def proposed_meta(topic: str, page_type: str, cluster: str, record: dict[str, Any]) -> str:
    current = compact(record.get("meta_description"))
    meta_phones = {
        clean_phone(value) for value in re.findall(r"(?:\+?91[\s-]?)?[6-9]\d(?:[\s-]?\d){8}", current)
    }
    phones_safe = not meta_phones or all(value in ALLOWED_PHONES for value in meta_phones)
    if 90 <= len(current) <= 165 and phones_safe and not re.search(r"\b(best|no\.?\s*1|world.class|unmatched)\b", current, re.I):
        return current
    if page_type == "Local SEO Page":
        location = location_from_topic(topic, record.get("canonical") or "")
        product = re.sub(r"\s+(?:in|near|for)\s+.+$", "", topic, flags=re.I)
        value = f"{product} for projects in {location}, with practical site-access, delivery and specification guidance. Share your requirement for a SAMAN quotation."
    elif page_type.startswith("Product"):
        value = f"Explore {topic} with published size, material, application, delivery and installation details from SAMAN Portable. Request a specification-based quotation."
    elif page_type == "Editorial / Guide":
        value = f"Use this practical {topic} guide to compare applications, specification factors, delivery requirements and buying questions before requesting a quote."
    elif page_type == "Homepage":
        value = "Explore SAMAN Portable's factory-built cabins, container spaces, prefab buildings and steel construction systems for projects across India."
    else:
        value = current or f"Practical information about {topic} from SAMAN Portable, including product scope, buyer guidance and enquiry details."
    return value[:165].rstrip(" ,;-")


def draft_index(root: Path) -> list[tuple[str, str, str]]:
    items: list[tuple[str, str, str]] = []
    for directory in [root / "page-structure/content-drafts", root / "content-drafts"]:
        if not directory.exists():
            continue
        for path in directory.glob("*.md"):
            try:
                text = path.read_text(encoding="utf-8", errors="ignore")
            except OSError:
                continue
            lowered = text.lower()
            state = "Approval/final marker found—verify governing packet" if re.search(r"\b(approved|final|owner-approved)\b", lowered) else "Draft/packet; approval verify"
            items.append((path.name, lowered, state))
    return items


def find_draft(url: str, drafts: list[tuple[str, str, str]]) -> str:
    path = urlsplit(url).path.lower().rstrip("/")
    slug = path.split("/")[-1]
    if len(slug) < 5:
        return ""
    hits = []
    needles = {path, f"/{slug}", slug.replace("-", " ")}
    for name, text, state in drafts:
        if any(needle in text or needle in name.lower() for needle in needles):
            hits.append(f"{state}: {name}")
        if len(hits) == 3:
            break
    return "; ".join(hits)


def indexability(record: dict[str, Any], url: str) -> str:
    if record.get("status") != 200:
        return "Not indexable (non-200)"
    if record.get("noindex") or re.search(r"\bnoindex\b", record.get("x_robots_tag", ""), re.I):
        return "Noindex"
    canonical = record.get("canonical", "")
    if canonical and canonical != url:
        return "Alternate canonical"
    return "Indexable"


def quality_score(url: str, record: dict[str, Any], page_type: str, sim: float, incoming: int) -> int:
    score = 100
    if record.get("status") != 200:
        return max(0, 30 if record.get("status") in {301, 308} else 10)
    title = compact(record.get("title"))
    meta = compact(record.get("meta_description"))
    if not record.get("canonical"):
        score -= 10
    elif record.get("canonical") != url:
        score -= 12
    if record.get("noindex"):
        score -= 15
    if not title:
        score -= 10
    elif not 25 <= len(title) <= 65:
        score -= 4
    if not meta:
        score -= 8
    elif not 70 <= len(meta) <= 170:
        score -= 3
    if record.get("h1_count") != 1:
        score -= 6
    wc = int(record.get("word_count") or 0)
    if wc < 250:
        score -= 18
    elif wc < 600:
        score -= 9
    if sim >= 0.80:
        score -= 25
    elif sim >= 0.60:
        score -= 14
    elif sim >= 0.40:
        score -= 6
    if page_type.startswith("Product") and not (record.get("technical_headings") or record.get("table_count")):
        score -= 8
    images = int(record.get("image_count") or 0)
    missing = int(record.get("missing_alt_count") or 0)
    if images and missing / images > 0.2:
        score -= 5
    if incoming == 0 and any(source.startswith("sitemap-") for source in record.get("sources", [])):
        score -= 5
    phone_values = {clean_phone(value) for value in record.get("phones", [])}
    if BANNED_PHONE in phone_values or any(value not in ALLOWED_PHONES for value in phone_values if len(value) == 12):
        score -= 8
    return max(0, min(100, score))


def sitemap_page(record: dict[str, Any]) -> bool:
    return any(source.startswith("sitemap-") and "images" not in source for source in record.get("sources", []))


def choose_exact_winners(records: dict[str, dict[str, Any]], gsc: dict[str, Any]) -> dict[str, str]:
    groups: dict[str, list[str]] = defaultdict(list)
    for url, record in records.items():
        if record.get("status") == 200 and record.get("text_hash"):
            groups[record["text_hash"]].append(url)
    winner_for: dict[str, str] = {}
    for urls in groups.values():
        if len(urls) < 2:
            continue
        def rank(url: str) -> tuple[int, float, int, int]:
            record = records[url]
            return (
                1 if sitemap_page(record) else 0,
                gsc.get(url, {}).get("Recent", {}).get("clicks", 0) + gsc.get(url, {}).get("16M", {}).get("clicks", 0),
                int(record.get("word_count") or 0),
                -len(url),
            )
        winner = max(urls, key=rank)
        for url in urls:
            winner_for[url] = winner
    return winner_for


def decision_for(
    url: str, record: dict[str, Any], page_type: str, cluster: str, sim: dict[str, Any],
    exact_winner: str, gsc_row: dict[str, Any], configured: dict[str, dict[str, Any]],
    all_gsc: dict[str, Any],
) -> tuple[str, str, str, str, str]:
    status = record.get("status", 0)
    recent_clicks = gsc_row.get("Recent", {}).get("clicks", 0)
    clicks_16 = gsc_row.get("16M", {}).get("clicks", 0)
    similarity = sim.get("similarity", 0)
    target = ""
    redirect_type = ""
    confidence = "High"
    reasons: list[str] = []

    if url in configured:
        row = configured[url]
        target = row.get("destination_url") or ""
        action = "KEEP EXISTING PERMANENT REDIRECT" if row.get("permanent") else "REVIEW TEMPORARY REDIRECT"
        redirect_type = str(row.get("status_code", ""))
        reasons.append("Configured redirect source; keep out of sitemap and internal links.")
        return action, target, redirect_type, "; ".join(reasons), confidence
    if status in {301, 302, 303, 307, 308}:
        target = normalize_url(record.get("location", "")) or ""
        action = "KEEP / DOCUMENT REDIRECT"
        redirect_type = str(status)
        reasons.append("Live redirect is not represented by a literal local rule or is pattern-driven.")
        return action, target, redirect_type, "; ".join(reasons), "Medium"
    if status in {404, 410}:
        if clicks_16 or recent_clicks:
            target = HUB_TARGETS.get(cluster, "")
            action = "301 IF CLOSE EQUIVALENT; OTHERWISE 410"
            reasons.append("Historical GSC equity exists; confirm intent/backlinks/leads before redirecting.")
            confidence = "Medium"
        else:
            action = "KEEP 410 / REMOVE INTERNAL REFERENCES"
            reasons.append("No performance evidence in capped local GSC exports; preserve Gone only if intentionally retired.")
            confidence = "Medium"
        return action, target, "301 or 410", "; ".join(reasons), confidence
    if status == 0:
        return "MANUAL FETCH / SERVER REVIEW", "", "", record.get("error", "Fetch failed"), "Low"
    if status != 200:
        return "MANUAL STATUS REVIEW", "", "", f"Unexpected HTTP {status}.", "Low"
    if exact_winner and exact_winner != url:
        return "301 MERGE EXACT DUPLICATE", exact_winner, "301/308", "Rendered main content is an exact duplicate of the selected canonical winner.", "High"
    canonical = record.get("canonical", "")
    if canonical and canonical != url:
        return "301 TO DECLARED CANONICAL", canonical, "301/308", "Page declares a different canonical; consolidate if the intents truly match.", "High"
    if record.get("noindex"):
        return "KEEP NOINDEX OR REMOVE FROM SITEMAP", "", "", "Noindex URL should not be advertised in an XML sitemap.", "High"
    wc = int(record.get("word_count") or 0)
    if page_type == "Local SEO Page":
        if similarity >= 0.80:
            if recent_clicks >= 3 or clicks_16 >= 15:
                return "REWRITE / NARROW LOCAL INTENT", "", "", "High similarity to another local page, but it has measurable search equity; differentiate before considering consolidation.", "High"
            target = HUB_TARGETS.get(cluster, "")
            return "301 DOORWAY-RISK LOCAL PAGE", target, "301/308", "High rendered-content similarity and no material GSC evidence in the available exports.", "High"
        if similarity >= 0.45 or wc < 650:
            return "KEEP ONLY AFTER LOCAL UNIQUENESS FIX", "", "", "Local page needs stronger city-specific buyer utility, logistics/site-access evidence and unique FAQs.", "Medium"
        return "KEEP & IMPROVE", "", "", "Distinct local content in the crawl; retain only while it serves a genuine product + location + intent need.", "Medium"
    if similarity >= 0.82:
        other = sim.get("nearest", "")
        other_clicks = all_gsc.get(other, {}).get("Recent", {}).get("clicks", 0) + all_gsc.get(other, {}).get("16M", {}).get("clicks", 0)
        own_clicks = recent_clicks + clicks_16
        if own_clicks <= other_clicks:
            return "301 MERGE NEAR DUPLICATE", other, "301/308", f"Near-duplicate content ({similarity:.0%}) with the stronger sibling.", "Medium"
    if page_type == "Editorial / Guide" and wc < 400 and not (recent_clicks or clicks_16):
        return "301 THIN GUIDE TO OWNER HUB", HUB_TARGETS.get(cluster, ""), "301/308", "Thin editorial page with no visibility in the capped GSC exports; validate backlinks/leads first.", "Medium"
    if page_type.startswith("Product") and wc < 500 and not (record.get("table_count") or record.get("technical_headings")):
        return "KEEP ONLY AFTER APPROVED REWRITE", "", "", "Commercial page is thin and lacks a comprehensive technical specification surface.", "High"
    return "KEEP & IMPROVE", "", "", "Unique canonical 200 URL; improve only the documented weak elements and preserve its single intent.", "High"


def issue_list(url: str, record: dict[str, Any], page_type: str, sim: dict[str, Any], incoming: int) -> str:
    issues = []
    title = compact(record.get("title"))
    meta = compact(record.get("meta_description"))
    if not title:
        issues.append("missing title")
    elif not 25 <= len(title) <= 65:
        issues.append(f"title length {len(title)}")
    if not meta:
        issues.append("missing meta description")
    elif not 70 <= len(meta) <= 170:
        issues.append(f"meta length {len(meta)}")
    if record.get("h1_count") != 1:
        issues.append(f"H1 count {record.get('h1_count', 0)}")
    if not record.get("canonical"):
        issues.append("missing canonical")
    elif record.get("canonical") != url:
        issues.append("non-self canonical")
    if page_type.startswith("Product") and not (record.get("technical_headings") or record.get("table_count")):
        issues.append("thin/missing technical specs")
    if sim.get("similarity", 0) >= 0.60:
        issues.append(f"near-duplicate {sim['similarity']:.0%}")
    if record.get("missing_alt_count"):
        issues.append(f"{record['missing_alt_count']} images missing alt")
    if incoming == 0 and sitemap_page(record):
        issues.append("orphan in crawl")
    phones = {clean_phone(value) for value in record.get("phones", [])}
    if BANNED_PHONE in phones:
        issues.append("banned phone present")
    if any(value not in ALLOWED_PHONES for value in phones if len(value) == 12):
        issues.append("non-approved phone present")
    return "; ".join(issues) or "No critical crawl defect; assess content usefulness and factual accuracy."


def add_table(ws, name: str) -> None:
    # Worksheet autofilters are used instead of OOXML table objects.  This
    # keeps the large audit sheets fully filterable while avoiding an Excel
    # interoperability defect seen with generated table definitions on this
    # Windows Office build.
    if ws.max_row >= 2 and ws.max_column >= 1:
        ws.auto_filter.ref = ws.dimensions


def style_sheet(ws, widths: dict[str, float] | None = None, freeze: str = "A2") -> None:
    ws.freeze_panes = freeze
    ws.sheet_view.showGridLines = False
    for cell in ws[1]:
        cell.font = Font(name="Arial", size=10, bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=GREEN)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=Side(style="thin", color=WHITE))
    ws.row_dimensions[1].height = 32
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="Arial", size=9, color=BLACK)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=THIN_GREY)
    if widths:
        for column, width in widths.items():
            ws.column_dimensions[column].width = width
    ws.auto_filter.ref = ws.dimensions


def write_rows(ws, headers: list[str], rows: Iterable[list[Any]]) -> None:
    ws.append(headers)
    for row in rows:
        ws.append(row)


def set_hyperlinks(ws, columns: Iterable[int], start_row: int = 2) -> None:
    for column in columns:
        for row in range(start_row, ws.max_row + 1):
            cell = ws.cell(row, column)
            if isinstance(cell.value, str) and cell.value.startswith("http"):
                cell.hyperlink = cell.value
                cell.style = "Hyperlink"
                cell.font = Font(name="Arial", size=9, color="0563C1", underline="single")


def parse_chart(path: Path) -> list[dict[str, Any]]:
    rows = []
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        for row in csv.DictReader(handle):
            rows.append({"date": row["Date"], "clicks": int(row["Clicks"]), "impressions": int(row["Impressions"]), "position": float(row["Position"])})
    return rows


def monthly_traffic(root: Path) -> list[dict[str, Any]]:
    base = parse_chart(root / "GSC_REPORTS/LAST_16_MONTHS/GSC_16M_Chart.csv")
    recent = parse_chart(root / "GSC_REPORTS/LAST_3_MONTHS/GSC_3M_Chart.csv")
    by_date = {row["date"]: row for row in base}
    by_date.update({row["date"]: row for row in recent})
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for key in sorted(by_date):
        grouped[key[:7]].append(by_date[key])
    result = []
    for month, rows in grouped.items():
        clicks = sum(row["clicks"] for row in rows)
        impressions = sum(row["impressions"] for row in rows)
        result.append({
            "month": month, "clicks": clicks, "impressions": impressions, "days": len(rows),
            "ctr": clicks / max(impressions, 1), "position": sum(row["position"] for row in rows) / len(rows),
        })
    return result


def main() -> None:
    root = Path(__file__).resolve().parents[2]
    out_dir = root / "outputs/full-site-seo-audit-2026-08-11"
    evidence_path = out_dir / "live-crawl-evidence.json"
    evidence = json.loads(evidence_path.read_text(encoding="utf-8"))
    records: dict[str, dict[str, Any]] = {
        url: record for url, record in evidence["records"].items()
        if not re.match(r"^/(?:_next|api)(?:/|$)", urlsplit(url).path)
    }
    gsc = load_gsc(root)
    redirects = load_redirects(root)
    redirect_rows = [row for row in redirects if row.get("source_url")]
    configured: dict[str, dict[str, Any]] = {}
    duplicate_redirect_rules: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in redirect_rows:
        duplicate_redirect_rules[row["source_url"]].append(row)
        configured[row["source_url"]] = row

    incoming: Counter[str] = Counter()
    for record in records.values():
        if record.get("status") == 200:
            incoming.update(record.get("internal_urls", []))
    similarities = similarity_map(records)
    exact_winners = choose_exact_winners(records, gsc)
    drafts = draft_index(root)

    all_rows: list[dict[str, Any]] = []
    for url in sorted(records):
        record = records[url]
        page_type, cluster = classify_page(url, record)
        sim = similarities.get(url, {"nearest": "", "similarity": 0.0})
        exact_winner = exact_winners.get(url, "")
        action, target, redirect_type, reason, confidence = decision_for(
            url, record, page_type, cluster, sim, exact_winner, gsc.get(url, {}), configured, gsc,
        )
        final_target, hops, loop = redirect_chain(url, records, configured)
        if target and hops > 1:
            target = final_target
            reason += f" Flatten configured/live chain to final target ({hops} hops detected)."
        topic = primary_topic(url, record, page_type)
        current_specs = "; ".join(record.get("technical_headings", [])[:5])
        tables = "; ".join(record.get("table_headers", [])[:5])
        if current_specs or tables:
            spec_status = f"Present: {record.get('table_count', 0)} table(s); {current_specs or tables}"
        else:
            spec_status = "No substantial spec heading/table detected in rendered HTML"
        if page_type.startswith("Product"):
            spec_action = f"Comprehensive table topics: {SPEC_TOPICS.get(cluster, 'identity; dimensions; materials; options; delivery; installation')}. Also include an approved indicative ex-GST supply-only price table, disclaimer and RFQ CTA. Use only approved/live verified values; draft required for gaps."
            if cluster == "Sandwich Panels":
                spec_action += " Use HSN 940690. Keep panel warranty on its own line: ‘5–10 years, confirmed at quotation’; never merge it with structural-frame warranty."
        elif page_type == "Local SEO Page":
            spec_action = "Do not duplicate a full product spec table. Add only location-relevant access, unloading, delivery and installation facts; link to the correct product owner page."
        else:
            spec_action = "Use specifications only where they answer the informational intent; source every numeric value from approved product data."
        if page_type == "Local SEO Page":
            content_topic = f"One product + one location + one intent: {topic}. Unique angle must be real local logistics/site conditions, not city-name substitution."
        elif page_type == "Editorial / Guide":
            content_topic = f"Informational support topic: {topic}. It should answer the question and support—not replace—the {cluster} commercial owner."
        else:
            content_topic = f"Primary topic: {topic}. Unique angle: {ANGLES.get(cluster, 'clear buyer intent, factual product boundary and practical decision guidance')}."
        score = quality_score(url, record, page_type, sim.get("similarity", 0), incoming[url])
        metrics_16 = gsc.get(url, {}).get("16M", {})
        metrics_recent = gsc.get(url, {}).get("Recent", {})
        target_record = records.get(target, {}) if target else {}
        sources = ", ".join(record.get("sources", []))
        phones = ", ".join(record.get("phones", []))
        all_rows.append({
            "priority": "P0" if any(term in action for term in ["DOORWAY", "EXACT DUPLICATE", "CHAIN", "MANUAL STATUS"]) or loop else ("P1" if action != "KEEP & IMPROVE" or score < 75 else "P2"),
            "confidence": confidence,
            "url": url,
            "sources": sources,
            "status": record.get("status", 0),
            "indexability": indexability(record, url),
            "page_type": page_type,
            "cluster": cluster,
            "topic": topic,
            "current_title": compact(record.get("title")),
            "current_meta": compact(record.get("meta_description")),
            "proposed_title": proposed_title(topic, page_type, cluster, record),
            "proposed_meta": proposed_meta(topic, page_type, cluster, record),
            "spec_status": spec_status,
            "spec_action": spec_action,
            "content_topic": content_topic,
            "action": action,
            "target": target,
            "redirect_type": redirect_type,
            "target_status": target_record.get("status", "") if target else "",
            "chain_hops": hops,
            "loop": "YES" if loop else "",
            "reason": reason,
            "issues": issue_list(url, record, page_type, sim, incoming[url]),
            "doorway_risk": "High" if page_type == "Local SEO Page" and sim.get("similarity", 0) >= 0.80 else ("Medium" if page_type == "Local SEO Page" and (sim.get("similarity", 0) >= 0.45 or record.get("word_count", 0) < 650) else "Low / not detected"),
            "duplicate_risk": "High" if exact_winner and exact_winner != url else ("High" if sim.get("similarity", 0) >= 0.80 else ("Medium" if sim.get("similarity", 0) >= 0.55 else "Low")),
            "nearest": sim.get("nearest", ""),
            "similarity": sim.get("similarity", 0),
            "clicks_16": metrics_16.get("clicks", 0),
            "impr_16": metrics_16.get("impressions", 0),
            "position_16": metrics_16.get("position", ""),
            "recent_clicks": metrics_recent.get("clicks", 0),
            "recent_impr": metrics_recent.get("impressions", 0),
            "recent_position": metrics_recent.get("position", ""),
            "score": score,
            "incoming": incoming[url],
            "outgoing": record.get("internal_link_count", 0),
            "canonical": record.get("canonical", ""),
            "robots": compact(f"{record.get('meta_robots', '')} {record.get('x_robots_tag', '')}"),
            "word_count": record.get("word_count", 0),
            "h1": compact(record.get("h1")),
            "h1_count": record.get("h1_count", 0),
            "h2_topics": "; ".join(record.get("h2s", [])[:12]),
            "table_headers": tables,
            "schema": ", ".join(record.get("schema_types", [])),
            "images": record.get("image_count", 0),
            "missing_alt": record.get("missing_alt_count", 0),
            "phones": phones,
            "draft": find_draft(url, drafts),
            "implementation": "Do not change technical values or page copy without an approved draft. Deploy only after owner approval and one-hop redirect QA.",
        })

    by_url = {row["url"]: row for row in all_rows}
    live_sitemap_rows = [row for row in all_rows if sitemap_page(records[row["url"]])]
    local_rows = [row for row in live_sitemap_rows if row["page_type"] == "Local SEO Page"]
    high_local = sum(1 for row in local_rows if row["doorway_risk"] == "High")
    medium_local = sum(1 for row in local_rows if row["doorway_risk"] == "Medium")
    avg_local_sim = sum(row["similarity"] for row in local_rows) / max(len(local_rows), 1)
    current_200 = [row for row in live_sitemap_rows if row["status"] == 200]
    duplicate_title_counts = Counter(row["current_title"].lower() for row in current_200 if row["current_title"])
    duplicate_meta_counts = Counter(row["current_meta"].lower() for row in current_200 if row["current_meta"])
    nonapproved_phone_pages = 0
    banned_phone_pages = 0
    for row in current_200:
        phone_values = {clean_phone(value) for value in records[row["url"]].get("phones", [])}
        if any(value not in ALLOWED_PHONES for value in phone_values if len(value) == 12):
            nonapproved_phone_pages += 1
        if BANNED_PHONE in phone_values:
            banned_phone_pages += 1

    wb = Workbook()
    wb.remove(wb.active)
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"

    # READ ME
    ws = wb.create_sheet("READ ME")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 110
    title = ws.cell(1, 1, "SAMAN Portable — Full-Site Google Policy & URL Remapping Audit")
    ws.merge_cells("A1:B1")
    title.font = Font(name="Arial", size=18, bold=True, color=WHITE)
    title.fill = PatternFill("solid", fgColor=GREEN)
    title.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 34
    readme = [
        ("Audit date", AUDIT_DATE.isoformat()),
        ("Outcome", "NOT certifiable as ‘100% Google compliant.’ Google provides no such certification, Core Updates are not a checklist, E-E-A-T is not a single score, and rankings are not guaranteed. This workbook identifies observable risks and implementation decisions."),
        ("Coverage", f"{len(records):,} normalized URLs fetched from live sitemaps, internal links, configured redirects and the locally available top-1,000 GSC page exports. Live sitemap pages reviewed: {len(live_sitemap_rows):,}. Configured literal redirect sources: {len(configured):,}."),
        ("Important GSC limit", "Each local page export is capped at 1,000 rows. ‘Absent from GSC’ means absent from that capped export—not zero impressions. Before any uncertain 301, check full Search Console API data, backlinks, leads, sales history and index state."),
        ("Decision meanings", "KEEP & IMPROVE = retain URL; KEEP ONLY AFTER… = retain only after the listed approved rewrite; 301 MERGE/TO… = retire source to closest equivalent; KEEP 410 = intentionally gone with no equivalent; MANUAL REVIEW = insufficient evidence."),
        ("Metadata rule", "Proposed titles/descriptions preserve acceptable current metadata and replace missing, overlong or hype-heavy metadata. They are implementation drafts, not proof Google will display them."),
        ("Technical-spec rule", "No unsupported dimensions, prices, warranties, ratings or material-performance values were invented. Cells list required specification topics and current rendered evidence; missing values require an approved content draft/source sheet."),
        ("Redirect safety", "A 301/308 must go directly to a live 200, self-canonical, indexable, same-intent URL. Remove retired sources from sitemaps, internal links, schema/feed and navigation. Never redirect everything to the homepage."),
        ("Owner workflow", "This is an audit/build artifact only. Page-copy changes require an approved draft. Deployment requires the owner’s typed authorization to Codex and post-deploy QA."),
        ("Google Helpful Content", GOOGLE_HELPFUL),
        ("Google Core Updates", GOOGLE_CORE),
        ("Google Spam Policies", GOOGLE_SPAM),
        ("Google Canonical Guidance", GOOGLE_CANONICAL),
        ("Google Page Experience", GOOGLE_PAGE_EXPERIENCE),
        ("Google Ranking Systems", GOOGLE_RANKING),
        ("Google Update History", GOOGLE_UPDATES),
    ]
    for index, (label, value) in enumerate(readme, 3):
        ws.cell(index, 1, label).font = Font(name="Arial", size=10, bold=True, color=GREEN)
        ws.cell(index, 2, value).font = Font(name="Arial", size=10)
        ws.cell(index, 2).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[index].height = 42 if index < 12 else 24
        if isinstance(value, str) and value.startswith("http"):
            ws.cell(index, 2).hyperlink = value
            ws.cell(index, 2).style = "Hyperlink"
    ws.freeze_panes = "A3"

    # Executive Summary
    ws = wb.create_sheet("Executive Summary")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 82
    ws["A1"] = "Executive Summary"
    ws.merge_cells("A1:C1")
    ws["A1"].font = Font(name="Arial", size=18, bold=True, color=WHITE)
    ws["A1"].fill = PatternFill("solid", fgColor=GREEN)
    ws["A1"].alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 34
    ws.append(["Metric", "Value", "Interpretation"])
    metrics = [
        ("Crawl inventory", len(records), "All normalized URLs found through the evidence sources."),
        ("Live sitemap pages", len(live_sitemap_rows), "Current page URLs, excluding image sitemap repetitions."),
        ("Live sitemap 200s", len(current_200), "Should be indexable, self-canonical and unique."),
        ("Local pages", len(local_rows), "Doorway risk assessed by rendered-content similarity and page utility signals."),
        ("High doorway-risk local pages", high_local, "Requires rewrite/narrowing or same-intent consolidation after full data checks."),
        ("Medium local uniqueness risk", medium_local, "Needs stronger local logistics/site evidence and unique buyer help."),
        ("Average nearest local 8-shingle similarity", avg_local_sim, "Diagnostic only; high similarity alone is not the entire doorway test."),
        ("Exact rendered duplicate losers", sum(1 for row in all_rows if "EXACT DUPLICATE" in row["action"]), "Use a single canonical winner and permanent one-hop redirect."),
        ("Configured literal redirect sources", len(configured), "Keep out of XML sitemaps and internal links."),
        ("Redirect chains/loops requiring review", sum(1 for row in all_rows if row["chain_hops"] > 1 or row["loop"]), "Flatten directly to final live winner; eliminate loops."),
        ("Current pages with duplicate titles", sum(1 for row in current_200 if duplicate_title_counts[row["current_title"].lower()] > 1), "Rewrite only where titles do not reflect a distinct intent."),
        ("Current pages with duplicate metas", sum(1 for row in current_200 if row["current_meta"] and duplicate_meta_counts[row["current_meta"].lower()] > 1), "Unique snippets improve clarity, but metadata alone cannot fix weak content."),
        ("Current images missing alt", sum(row["missing_alt"] for row in current_200), "Repair using accurate image descriptions; do not stuff keywords."),
        ("Current pages with non-approved phone numbers", nonapproved_phone_pages, "Replace with South +91 88616 22859 or North +91 87960 39938 according to the approved CTA rule."),
        ("Current pages containing banned +91 62009 09435", banned_phone_pages, "Must remain zero site-wide."),
    ]
    for metric, value, note in metrics:
        ws.append([metric, value, note])
    for row in range(3, ws.max_row + 1):
        ws.cell(row, 1).font = Font(name="Arial", size=10, bold=True, color=GREEN)
        ws.cell(row, 2).font = Font(name="Arial", size=10, bold=True)
        ws.cell(row, 3).font = Font(name="Arial", size=10)
        ws.cell(row, 3).alignment = Alignment(wrap_text=True, vertical="top")
    ws["B9"].number_format = "0.0%"
    decision_header = ws.max_row + 2
    ws.cell(decision_header, 1, "URL decision counts")
    ws.cell(decision_header, 1).font = Font(name="Arial", size=12, bold=True, color=WHITE)
    ws.cell(decision_header, 1).fill = PatternFill("solid", fgColor=GREEN)
    ws.merge_cells(start_row=decision_header, start_column=1, end_row=decision_header, end_column=3)
    decision_formulas = [
        ("Keep & improve", '=COUNTIF(\'All URLs\'!$S:$S,"KEEP & IMPROVE")', "Live/historical canonical pages retained."),
        ("Keep only after approved fix", '=COUNTIF(\'All URLs\'!$S:$S,"KEEP ONLY*")', "Must not remain indexable in the same weak state."),
        ("Existing permanent redirects", '=COUNTIF(\'All URLs\'!$S:$S,"KEEP EXISTING PERMANENT REDIRECT")', "Already retired sources; keep one hop and out of internal signals."),
        ("New/review 301 decisions", '=COUNTIF(\'All URLs\'!$S:$S,"*301*")', "Require full-data, same-intent and owner gates."),
        ("Manual review", '=COUNTIF(\'All URLs\'!$S:$S,"MANUAL*")', "Insufficient or unexpected technical evidence."),
    ]
    for label, formula, note in decision_formulas:
        decision_header += 1
        ws.cell(decision_header, 1, label).font = Font(name="Arial", size=10, bold=True, color=GREEN)
        ws.cell(decision_header, 2, formula).font = Font(name="Arial", size=10, bold=True)
        ws.cell(decision_header, 3, note).font = Font(name="Arial", size=10)
        ws.cell(decision_header, 3).alignment = Alignment(wrap_text=True, vertical="top")
    start = ws.max_row + 3
    ws.cell(start, 1, "Recovery diagnosis")
    ws.cell(start, 1).font = Font(name="Arial", size=13, bold=True, color=WHITE)
    ws.cell(start, 1).fill = PatternFill("solid", fgColor=GREEN)
    ws.merge_cells(start_row=start, start_column=1, end_row=start, end_column=3)
    findings = [
        "The local 16-month GSC chart peaks around 12.4k clicks/month, not 55k; the earlier 55k claim requires GA or older GSC evidence not present here.",
        "Clicks dropped sharply during the 26 Aug–22 Sep 2025 spam update window (correlation, not proof of a penalty). They declined again around the Dec 2025 and Mar 2026 core/spam periods.",
        "Recovery began in May–July 2026. Do not bulk-delete pages that are already improving; consolidate only where intent equivalence, quality, links and business value support it.",
        "The strongest observable site risks are URL sprawl/legacy redirects, overlapping product terminology, inconsistent commercial facts across pages, and pages whose specification/trust surfaces are thin or contradictory.",
        "Manual actions, security issues, backlink schemes, paid-link history, user-agent cloaking and full index coverage cannot be certified from a public crawl. Check the corresponding Search Console reports before rollout.",
    ]
    for finding in findings:
        start += 1
        ws.cell(start, 1, "• " + finding)
        ws.merge_cells(start_row=start, start_column=1, end_row=start, end_column=3)
        ws.cell(start, 1).font = Font(name="Arial", size=10)
        ws.cell(start, 1).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[start].height = 32
    ws.freeze_panes = "A3"

    # All URLs
    ws = wb.create_sheet("All URLs")
    all_headers = [
        "Priority", "Confidence", "URL", "Inventory Sources", "HTTP", "Indexability", "Page Type", "Cluster", "Primary Topic / Intent",
        "Current Meta Title", "Current Meta Description", "Proposed Meta Title", "Title Len", "Proposed Meta Description", "Meta Len",
        "Technical Spec Coverage", "Technical Spec / Content Requirement", "Content Topic / Unique Angle", "Final Action", "Redirect Target",
        "Redirect Type", "Target HTTP", "Chain Hops", "Loop", "Decision Evidence", "Observed Issues", "Doorway Risk", "Duplicate Risk",
        "Nearest Similar URL", "Similarity", "16M Clicks", "16M Impressions", "16M Position", "Recent Clicks", "Recent Impressions", "Recent Position",
        "Diagnostic Score /100", "Incoming Internal Links", "Outgoing Internal Links", "Canonical", "Robots", "Word Count", "H1", "H1 Count",
        "H2 Topics", "Table Headers", "Schema Types", "Images", "Missing Alt", "Phones Found", "Draft Evidence", "Implementation Guardrail",
    ]
    ws.append(all_headers)
    for row_index, row in enumerate(all_rows, 2):
        values = [
            row["priority"], row["confidence"], row["url"], row["sources"], row["status"], row["indexability"], row["page_type"], row["cluster"], row["topic"],
            row["current_title"], row["current_meta"], row["proposed_title"], f"=LEN(L{row_index})", row["proposed_meta"], f"=LEN(N{row_index})",
            row["spec_status"], row["spec_action"], row["content_topic"], row["action"], row["target"], row["redirect_type"], row["target_status"], row["chain_hops"], row["loop"],
            row["reason"], row["issues"], row["doorway_risk"], row["duplicate_risk"], row["nearest"], row["similarity"], row["clicks_16"], row["impr_16"], row["position_16"],
            row["recent_clicks"], row["recent_impr"], row["recent_position"], row["score"], row["incoming"], row["outgoing"], row["canonical"], row["robots"], row["word_count"],
            row["h1"], row["h1_count"], row["h2_topics"], row["table_headers"], row["schema"], row["images"], row["missing_alt"], row["phones"], row["draft"], row["implementation"],
        ]
        ws.append(values)
    style_sheet(ws, {
        "A": 9, "B": 10, "C": 54, "D": 24, "E": 8, "F": 20, "G": 19, "H": 22, "I": 32, "J": 38, "K": 52, "L": 38, "M": 9, "N": 58, "O": 9,
        "P": 42, "Q": 58, "R": 58, "S": 34, "T": 52, "U": 12, "V": 10, "W": 10, "X": 8, "Y": 55, "Z": 48, "AA": 14, "AB": 14,
        "AC": 54, "AD": 11, "AE": 11, "AF": 14, "AG": 11, "AH": 13, "AI": 13, "AJ": 12, "AK": 12, "AL": 13, "AM": 11, "AN": 14, "AO": 14,
        "AP": 14, "AQ": 12, "AR": 12, "AS": 52, "AT": 14, "AU": 24, "AV": 10, "AW": 10, "AX": 30, "AY": 55, "AZ": 60,
    })
    set_hyperlinks(ws, [3, 20, 29, 40])
    for row in range(2, ws.max_row + 1):
        ws.cell(row, 30).number_format = "0.0%"
        if ws.cell(row, 1).value == "P0":
            ws.cell(row, 1).fill = PatternFill("solid", fgColor=RED_FILL)
        elif ws.cell(row, 1).value == "P1":
            ws.cell(row, 1).fill = PatternFill("solid", fgColor=AMBER_FILL)
    ws.conditional_formatting.add(f"AK2:AK{ws.max_row}", CellIsRule(operator="lessThan", formula=["60"], fill=PatternFill("solid", fgColor=RED_FILL)))
    add_table(ws, "AllURLsTable")

    # Keep & Improve
    ws = wb.create_sheet("Keep & Improve")
    keep_headers = [
        "Priority", "URL", "Page Type", "Cluster", "Primary Topic", "Final Action", "Proposed Meta Title", "Title Len", "Proposed Meta Description", "Meta Len",
        "Technical Spec Coverage", "Required Technical Spec Topics", "Content Topic / Unique Angle", "Main Issues", "16M Clicks", "Recent Clicks", "Position",
        "Diagnostic Score", "Draft Evidence", "Owner / Draft Gate", "Implementation Status",
    ]
    ws.append(keep_headers)
    keep_rows = [row for row in live_sitemap_rows if not any(token in row["action"] for token in ["301 ", "REDIRECT", "410", "MANUAL STATUS"])]
    for row_index, row in enumerate(keep_rows, 2):
        ws.append([
            row["priority"], row["url"], row["page_type"], row["cluster"], row["topic"], row["action"], row["proposed_title"], f"=LEN(G{row_index})",
            row["proposed_meta"], f"=LEN(I{row_index})", row["spec_status"], row["spec_action"], row["content_topic"], row["issues"], row["clicks_16"],
            row["recent_clicks"], row["recent_position"], row["score"], row["draft"], "Approved draft required before page-copy/spec changes", "Planned",
        ])
    style_sheet(ws, {"A": 9, "B": 54, "C": 20, "D": 22, "E": 34, "F": 28, "G": 40, "H": 9, "I": 60, "J": 9, "K": 42, "L": 62, "M": 62, "N": 58, "O": 11, "P": 11, "Q": 11, "R": 12, "S": 48, "T": 35, "U": 18})
    set_hyperlinks(ws, [2])
    dv = DataValidation(type="list", formula1='"Planned,Draft Needed,Approved,Implemented,QA Passed"', allow_blank=False)
    ws.add_data_validation(dv)
    dv.add(f"U2:U{ws.max_row}")
    add_table(ws, "KeepImproveTable")

    # Redirect Map
    ws = wb.create_sheet("Redirect Map")
    redirect_headers = [
        "Priority", "Source URL", "Live HTTP", "Current / Recommended Target", "Final Target", "Redirect Code", "Target HTTP", "Chain Hops", "Loop",
        "Action", "Cluster", "Reason", "16M Clicks", "Recent Clicks", "Source in Sitemap", "Source Internal Links", "Target Self-Canonical", "Target Indexable",
        "Confidence", "Pre-Deploy Gate", "Implementation Status", "Post-Deploy Result",
    ]
    ws.append(redirect_headers)
    redirect_decisions = [row for row in all_rows if row["target"] or any(token in row["action"] for token in ["REDIRECT", "301", "410"])]
    for row in redirect_decisions:
        final_target, hops, loop = redirect_chain(row["url"], records, configured)
        target = row["target"] or (final_target if final_target != row["url"] else "")
        target_record = records.get(target, {})
        source_in_sitemap = "YES" if sitemap_page(records[row["url"]]) else ""
        target_self = "YES" if target and target_record.get("canonical") == target else ("NO" if target else "")
        target_indexable = "YES" if target and indexability(target_record, target) == "Indexable" else ("NO" if target else "")
        ws.append([
            row["priority"], row["url"], row["status"], row["target"], final_target if final_target != row["url"] else target, row["redirect_type"],
            target_record.get("status", ""), hops, "YES" if loop else "", row["action"], row["cluster"], row["reason"], row["clicks_16"], row["recent_clicks"],
            source_in_sitemap, row["incoming"], target_self, target_indexable, row["confidence"], "Check full GSC, backlinks, leads/sales, same intent, target 200/self-canonical; owner approve", "Planned", "",
        ])
    style_sheet(ws, {"A": 9, "B": 55, "C": 10, "D": 55, "E": 55, "F": 12, "G": 11, "H": 11, "I": 8, "J": 34, "K": 22, "L": 64, "M": 11, "N": 11, "O": 14, "P": 14, "Q": 16, "R": 15, "S": 11, "T": 68, "U": 18, "V": 28})
    set_hyperlinks(ws, [2, 4, 5])
    dv = DataValidation(type="list", formula1='"Planned,Validated,Owner Approved,Implemented,QA Passed,Rejected"', allow_blank=False)
    ws.add_data_validation(dv)
    dv.add(f"U2:U{ws.max_row}")
    for row in range(2, ws.max_row + 1):
        if ws.cell(row, 8).value and ws.cell(row, 8).value > 1:
            ws.cell(row, 8).fill = PatternFill("solid", fgColor=RED_FILL)
        if ws.cell(row, 15).value == "YES":
            ws.cell(row, 15).fill = PatternFill("solid", fgColor=RED_FILL)
    add_table(ws, "RedirectMapTable")

    # Doorway audit
    ws = wb.create_sheet("Doorway & Local Pages")
    doorway_headers = ["URL", "Cluster", "Primary Local Intent", "Word Count", "Nearest Local URL", "8-Word Shingle Similarity", "Location Mentions (approx.)", "16M Clicks", "Recent Clicks", "Doorway Risk", "Decision", "Required Local Proof / Fix", "Notes"]
    ws.append(doorway_headers)
    for row in local_rows:
        location = location_from_topic(row["topic"], row["url"])
        content = records[row["url"]].get("content_text", "").lower()
        mentions = content.count(location.lower()) if location != "the location" else 0
        notes = "Google doorway test also asks whether the page is merely an entry/funnel and whether a useful browseable hierarchy exists; similarity is diagnostic, not a verdict by itself."
        ws.append([row["url"], row["cluster"], row["topic"], row["word_count"], row["nearest"], row["similarity"], mentions, row["clicks_16"], row["recent_clicks"], row["doorway_risk"], row["action"], "Real local demand/use cases; route/access/unloading; climate/site constraints where true; unique FAQs; correct hub link; no fake office/project/client", notes])
    style_sheet(ws, {"A": 55, "B": 22, "C": 38, "D": 12, "E": 55, "F": 14, "G": 15, "H": 11, "I": 11, "J": 14, "K": 32, "L": 70, "M": 70})
    set_hyperlinks(ws, [1, 5])
    for row in range(2, ws.max_row + 1):
        ws.cell(row, 6).number_format = "0.0%"
    add_table(ws, "DoorwayLocalTable")

    # Duplicates & cannibalization
    ws = wb.create_sheet("Duplicates & Cannibalization")
    dup_headers = ["URL A", "URL B / Winner", "Relationship", "Similarity", "Cluster", "Page Type", "A 16M Clicks", "B 16M Clicks", "A Action", "Evidence / Boundary Note"]
    ws.append(dup_headers)
    emitted: set[tuple[str, str, str]] = set()
    for row in live_sitemap_rows:
        if row["nearest"] and row["similarity"] >= 0.35:
            key = tuple(sorted([row["url"], row["nearest"]])) + ("near",)
            if key not in emitted:
                emitted.add(key)
                other = by_url.get(row["nearest"], {})
                ws.append([row["url"], row["nearest"], "Rendered near-duplicate", row["similarity"], row["cluster"], row["page_type"], row["clicks_16"], other.get("clicks_16", 0), row["action"], "Merge only if intent is equivalent; otherwise differentiate H1/title/intro/specs/FAQs/internal anchors."])
        if exact_winners.get(row["url"]) and exact_winners[row["url"]] != row["url"]:
            key = (row["url"], exact_winners[row["url"]], "exact")
            if key not in emitted:
                emitted.add(key)
                winner = by_url.get(exact_winners[row["url"]], {})
                ws.append([row["url"], exact_winners[row["url"]], "Exact rendered duplicate", 1.0, row["cluster"], row["page_type"], row["clicks_16"], winner.get("clicks_16", 0), row["action"], "Use one canonical winner; repoint internal links and consolidate signals."])
    risk_pairs = [
        ("Porta Cabins", "Portable Cabin", "Highest: construction-market porta-cabin category vs general relocatable portable cabin."),
        ("PEB Constructions", "Pre-Engineered Buildings", "Highest: execution service vs engineered building system/product."),
        ("Portable Office", "Container Offices", "Medium: office-focused modular workspace vs heavy-duty container-based office."),
        ("Container Houses", "Prefabricated Houses", "Medium: container-based residence vs broader prefab residential system."),
        ("Prefab Buildings", "Prefabricated Houses", "Medium: commercial/institutional category vs residential homes."),
    ]
    for a, b, note in risk_pairs:
        ws.append([a, b, "Standing cluster boundary", "", f"{a} ↔ {b}", "Sitewide", "", "", "Protect both owners", note])
    style_sheet(ws, {"A": 55, "B": 55, "C": 28, "D": 12, "E": 28, "F": 20, "G": 12, "H": 12, "I": 32, "J": 78})
    set_hyperlinks(ws, [1, 2])
    for row in range(2, ws.max_row + 1):
        ws.cell(row, 4).number_format = "0.0%"
    add_table(ws, "DuplicateCannibalTable")

    # Policy matrix
    ws = wb.create_sheet("Google Policy Matrix")
    policy_headers = ["Area", "Google Requirement (plain language)", "Audit Status", "Observed Evidence", "Required Action", "Verification Limit", "Official Source"]
    policy_rows = [
        ["Helpful / people-first content", "Original, substantial, accurate content that satisfies an intended audience and demonstrates first-hand expertise.", "NEEDS IMPROVEMENT", "Many strong product pages exist, but thin/overlapping pages and sitewide commercial-fact contradictions remain.", "Fix factual truth sheet; improve only documented weak pages; add real project/manufacturing proof.", "Helpfulness is assessed by multiple ranking signals; no public compliance score exists.", GOOGLE_HELPFUL],
        ["Core Updates", "Assess impacted pages/queries objectively; avoid quick-fix churn and delete only as a last resort.", "RECOVERY IN PROGRESS", "GSC clicks fell sharply in Aug–Sep 2025 and bottomed in Apr 2026, then improved in Jun–Jul 2026.", "Prioritize high-equity pages and verified conflicts; monitor after changes through future updates.", "Timing correlation does not prove an algorithmic cause or penalty.", GOOGLE_CORE],
        ["Doorway abuse", "Do not create city/query pages that are substantially similar entry pages or funnels to the same final destination.", "HIGH RISK" if high_local else ("WATCH" if medium_local else "NO HIGH-RISK PATTERN DETECTED"), f"{len(local_rows)} local pages; {high_local} high-risk and {medium_local} medium-risk by crawl diagnostics; average nearest 8-shingle similarity {avg_local_sim:.1%}.", "Use the dedicated local-page sheet. Keep genuinely useful local pages; rewrite/narrow or consolidate only evidence-backed risks.", "A public crawl cannot determine creator intent; manual review of local utility is required.", GOOGLE_SPAM],
        ["Scaled content abuse", "Large volumes of unoriginal/low-value pages made mainly to manipulate ranking violate policy, regardless of creation method.", "WATCH / REMEDIATE IDENTIFIED DUPLICATES", f"{len(live_sitemap_rows)} sitemap pages plus a large historical URL/redirect footprint; exact and near duplicates are listed separately.", "Stop search-variation page creation; merge equivalent intents; require evidence and unique buyer value.", "Authorship/automation history is not fully observable.", GOOGLE_SPAM],
        ["Keyword stuffing", "Avoid unnatural repetition, city lists and repeated phone/keyword blocks.", "PAGE-LEVEL REVIEW NEEDED", "Metadata/H1/phone inconsistencies are surfaced per URL; crawl does not infer intent from frequency alone.", "Rewrite only unnatural blocks; keep terminology natural and cluster-owned.", "Frequency is not itself a violation; human reading is needed.", GOOGLE_SPAM],
        ["Cloaking", "Do not show materially different content to crawlers and users to manipulate ranking.", "NOT DETECTED; NOT CERTIFIED", "The audit used one normal browser-like user agent and found rendered HTML.", "Compare Googlebot/mobile and browser renders; inspect Search Console URL Inspection.", "Cannot verify user-agent/device variants from a single crawl.", GOOGLE_SPAM],
        ["Hidden text/link abuse", "Do not hide text/links solely for ranking manipulation.", "NOT FULLY VERIFIABLE", "No intentional hidden-spam conclusion from HTML alone.", "Review computed CSS/DOM and template blocks; accordion/tab content is allowed when user-accessible.", "Requires browser/CSS and intent review.", GOOGLE_SPAM],
        ["Sneaky redirects", "Redirects must not deceive; legitimate consolidation is allowed.", "TECHNICAL CLEANUP NEEDED", "Configured sources, live targets, chains and loops are mapped in Redirect Map.", "Use same-intent one-hop 301/308s; never send unrelated pages to home.", "Conditional redirects may require device/referrer tests.", GOOGLE_SPAM],
        ["Canonical / duplicate URLs", "Select one preferred URL; redirects and canonical tags are strong signals, sitemap inclusion is weaker.", "NEEDS CLEANUP", "Canonical conflicts, sitemap redirect sources and duplicate rendered content are listed per URL.", "Align redirect, self-canonical, sitemap and internal links on the same winner.", "Google may select a different canonical.", GOOGLE_CANONICAL],
        ["Link spam", "Do not buy/sell/manipulate ranking links or use automated/low-quality link schemes.", "NOT VERIFIABLE FROM SITE CRAWL", "Internal links were inventoried; backlink acquisition history was not provided.", "Audit paid links/backlinks; qualify sponsored links; avoid exact-match network links.", "Requires Search Console links/manual actions and link-source evidence.", GOOGLE_SPAM],
        ["Hacked/malicious content", "Prevent unauthorized injected, deceptive or harmful content.", "NO ISSUE DETECTED; CHECK GSC", "No obvious hacked URLs in discovered inventory.", "Check Security Issues, server logs and malware scans.", "Public crawl cannot certify absence of conditional injection.", GOOGLE_SPAM],
        ["Scraping / thin affiliation", "Do not republish third-party content or merchant descriptions without substantial original value.", "NO AFFILIATE PATTERN DETECTED", "Site is primarily first-party product/service content.", "Retain original manufacturing/buyer evidence and source external facts.", "Copyright/source provenance not fully tested.", GOOGLE_SPAM],
        ["Site reputation abuse", "Do not host unrelated third-party content mainly to exploit the host’s ranking signals.", "NO OBVIOUS PATTERN DETECTED", "Discovered content remains within portable/prefab/steel-building themes.", "Keep strict topical/editorial ownership and review third-party submissions.", "Employment/third-party commissioning relationships were not provided.", GOOGLE_SPAM],
        ["User-generated spam", "Control spam in comments/reviews/uploads.", "CHECK OPERATIONAL CONTROLS", "Review surfaces exist; public crawl does not verify moderation/wiring.", "Moderate reviews/comments, use anti-spam controls, and publish only genuine reviews.", "Requires admin access and moderation logs.", GOOGLE_SPAM],
        ["Page experience / CWV", "Provide good CWV, HTTPS, mobile display, non-intrusive UX and clearly distinguish main content.", "SEPARATE CWV WORKSTREAM", "Workspace has prior mobile CWV reports; this workbook focuses URL/content policy and remapping.", "Meet LCP ≤2.5s and near-zero CLS on mobile; re-test representative templates after changes.", "CWV alone does not guarantee ranking.", GOOGLE_PAGE_EXPERIENCE],
        ["Manual actions / Security Issues", "Resolve any Search Console actions/issues before expecting full eligibility.", "MUST CHECK IN SEARCH CONSOLE", "No authenticated Manual Actions or Security Issues export was provided.", "Owner checks both reports and records screenshots/date before implementation.", "Impossible to certify publicly.", GOOGLE_SPAM],
    ]
    write_rows(ws, policy_headers, policy_rows)
    style_sheet(ws, {"A": 28, "B": 62, "C": 26, "D": 70, "E": 70, "F": 58, "G": 58})
    set_hyperlinks(ws, [7])
    add_table(ws, "GooglePolicyTable")

    # Traffic recovery
    ws = wb.create_sheet("Traffic Recovery")
    traffic_headers = ["Month", "Clicks", "Impressions", "Days in Export", "CTR", "Average Position", "Projected Full-Month Clicks", "Notes"]
    ws.append(traffic_headers)
    traffic = monthly_traffic(root)
    for row_index, row in enumerate(traffic, 2):
        note = "Partial month; projection is directional only" if row["days"] < 27 else "Complete/near-complete month"
        ws.append([row["month"], row["clicks"], row["impressions"], row["days"], row["ctr"], row["position"], f"=IF(D{row_index}=0,0,B{row_index}/D{row_index}*DAY(EOMONTH(DATE(LEFT(A{row_index},4),RIGHT(A{row_index},2),1),0)))", note])
    style_sheet(ws, {"A": 12, "B": 12, "C": 16, "D": 14, "E": 12, "F": 16, "G": 22, "H": 42})
    for row in range(2, ws.max_row + 1):
        ws.cell(row, 5).number_format = "0.0%"
        ws.cell(row, 6).number_format = "0.0"
        ws.cell(row, 7).number_format = "0"
    chart = LineChart()
    chart.title = "Google Search Click Recovery"
    chart.y_axis.title = "Clicks"
    chart.x_axis.title = "Month"
    data = Reference(ws, min_col=2, min_row=1, max_row=ws.max_row)
    cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 8
    chart.width = 15
    ws.add_chart(chart, "J2")
    incident_start = ws.max_row + 3
    ws.cell(incident_start, 1, "Official Google ranking updates (correlation markers)")
    ws.cell(incident_start, 1).font = Font(name="Arial", size=12, bold=True, color=WHITE)
    ws.cell(incident_start, 1).fill = PatternFill("solid", fgColor=GREEN)
    ws.merge_cells(start_row=incident_start, start_column=1, end_row=incident_start, end_column=8)
    incidents = [
        ("2025-03-13", "March 2025 core update"), ("2025-06-30", "June 2025 core update"),
        ("2025-08-26", "August 2025 spam update"), ("2025-12-11", "December 2025 core update"),
        ("2026-03-24", "March 2026 spam update"), ("2026-03-27", "March 2026 core update"),
        ("2026-05-21", "May 2026 core update"), ("2026-06-24", "June 2026 spam update"),
    ]
    for event_date, name in incidents:
        incident_start += 1
        ws.cell(incident_start, 1, event_date)
        ws.cell(incident_start, 2, name)
        ws.cell(incident_start, 3, "Correlation is not causation; use URL/query comparison before and after completed rollouts.")
        ws.cell(incident_start, 8, GOOGLE_UPDATES)
        ws.cell(incident_start, 8).hyperlink = GOOGLE_UPDATES
        ws.cell(incident_start, 8).style = "Hyperlink"

    # Redirect rule conflicts
    ws = wb.create_sheet("Redirect Rule QA")
    qa_headers = ["Source URL", "Rule Count", "Destinations", "Status Codes", "Risk", "Required Fix"]
    ws.append(qa_headers)
    for source, rows in sorted(duplicate_redirect_rules.items()):
        destinations = sorted(set(row.get("destination_url") or row.get("destination_pattern") for row in rows))
        codes = sorted(set(row.get("status_code") for row in rows))
        risk = "CRITICAL conflicting targets" if len(destinations) > 1 else ("Duplicate same target" if len(rows) > 1 else "Single rule")
        ws.append([source, len(rows), "; ".join(destinations), ", ".join(str(code) for code in codes), risk, "Keep one authoritative rule; verify final target is 200/self-canonical and route is one hop." if len(rows) > 1 else "No duplicate-rule action."])
    style_sheet(ws, {"A": 55, "B": 12, "C": 70, "D": 16, "E": 28, "F": 70})
    set_hyperlinks(ws, [1])
    add_table(ws, "RedirectRuleQATable")

    # Implementation checklist
    ws = wb.create_sheet("Implementation QA")
    checklist_headers = ["Gate", "Required Check", "Why It Prevents Mistakes", "Owner", "Status", "Evidence / Link"]
    checklist = [
        ("0", "Owner approves this exact workbook version and decision rows.", "Prevents applying stale or unapproved remaps.", "Owner", "Not Started", ""),
        ("1", "Export complete GSC page/query data beyond the local 1,000-row cap.", "Avoids treating ‘not in export’ as zero value.", "SEO", "Not Started", ""),
        ("2", "Check backlinks, leads, sales and CRM history for every proposed new 301/410.", "Prevents destroying off-site equity or business value.", "SEO + Sales", "Not Started", ""),
        ("3", "Confirm one page = one keyword = one intent and product boundary.", "Prevents new cannibalization.", "Content", "Not Started", ""),
        ("4", "Confirm every target is live 200, self-canonical, indexable and same-intent.", "Prevents soft 404s, unrelated redirects and signal loss.", "Developer", "Not Started", ""),
        ("5", "Flatten each redirect to one permanent server hop (301 or 308).", "Avoids redirect chains and crawl waste.", "Developer", "Not Started", ""),
        ("6", "Remove retired sources from XML sitemaps, internal links, JSON-LD, feeds and menus.", "Aligns every canonical signal on the winner.", "Developer", "Not Started", ""),
        ("7", "Do not change technical facts, prices, warranty, SKU or CTA without approved source draft.", "Enforces zero invention and commercial truth.", "Content + Owner", "Not Started", ""),
        ("8", "Run staging crawl: zero loops; zero chains; zero sitemap redirects/noindex; zero broken internal links.", "Catches remapping errors before release.", "QA", "Not Started", ""),
        ("9", "Owner gives typed deploy authorization to Codex.", "Required deployment gate.", "Owner", "Not Started", ""),
        ("10", "Post-deploy crawl and compare all source/target status, canonical and sitemap states.", "Confirms live behavior matches workbook.", "QA", "Not Started", ""),
        ("11", "Monitor GSC indexing, clicks, impressions and query ownership at 7/28/60/90 days.", "Recovery and redirect absorption take time.", "SEO", "Not Started", ""),
        ("12", "Check Search Console Manual Actions and Security Issues.", "Public crawling cannot certify these reports.", "Owner", "Not Started", ""),
    ]
    write_rows(ws, checklist_headers, checklist)
    style_sheet(ws, {"A": 9, "B": 72, "C": 62, "D": 18, "E": 18, "F": 55})
    dv = DataValidation(type="list", formula1='"Not Started,In Progress,Passed,Blocked,Not Applicable"', allow_blank=False)
    ws.add_data_validation(dv)
    dv.add(f"E2:E{ws.max_row}")
    add_table(ws, "ImplementationQATable")

    # Add comments and workbook-wide number formats.
    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.font.name is None:
                    cell.font = Font(name="Arial", size=9)
    wb["All URLs"]["AK1"].comment = Comment("Diagnostic score based on observable crawl signals. It is not a Google score, ranking factor or compliance certificate.", "Codex")
    wb["Doorway & Local Pages"]["F1"].comment = Comment("Jaccard similarity of rendered main-content 8-word shingles. Use with utility, funnel behavior, intent and local evidence—not alone.", "Codex")

    output_path = out_dir / "SAMAN_Full_Site_Google_Policy_URL_Remapping_Audit_2026-08-11.xlsx"
    wb.save(output_path)
    print(json.dumps({
        "output": str(output_path), "sheets": wb.sheetnames, "all_urls": len(all_rows),
        "live_sitemap_urls": len(live_sitemap_rows), "keep_rows": len(keep_rows),
        "redirect_rows": len(redirect_decisions), "local_rows": len(local_rows),
        "high_local_risk": high_local, "medium_local_risk": medium_local,
    }, indent=2))


if __name__ == "__main__":
    main()

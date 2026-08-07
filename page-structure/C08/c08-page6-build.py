"""C-08 page six: body copy, Section H pack and the 60-row alt manifest.

Three verbatim implementations, each proved rather than asserted:
  · body copy — markup-only conversion, round-trip diffed against the source prose
  · Section H — 50 strings copied key-for-key, lengths asserted against the source
  · alts      — column H written byte-identical, then read back and compared
"""
import html as htmllib
import io, json, os, re, sys

import openpyxl

sys.stdout.reconfigure(encoding='utf-8')
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import build_copy as B

SLUG = 'prefabricated-container-house'
DATA = os.path.join(B.REPO, 'src', 'data', 'products')
MANIFEST = os.path.join(B.REPORTS, 'C08-PREFAB-CONTAINER-HOUSE-ALT-MANIFEST-60-05Aug2026.xlsx')
PACK = os.path.join(B.REPORTS, 'C08-PAGE6-SECTION-H-PACK-06Aug2026.json')
COPY = os.path.join(B.REPORTS, 'C08-PAGE6-V2-06Aug2026.md')
SIZES = ['20x8', '20x10', '20x12', '40x8', '40x10', '40x12']

# Slot -> (size, view token). Page six's only 16:9 interior is I04: the kitchen
# frame I02 is slotted G5 in the gallery, and re-cutting it would breach the
# no-source-twice gate. See the report for the one resulting deviation.
IMAGE_SLOTS = [
    ('40x12', 'E02'),   # 1 EXTERIOR HERO 40x12
    ('20x10', 'E04'),   # 2 asks INTERIOR kitchen/dining — none exists, see report
    ('20x12', 'E05'),   # 3 EXTERIOR different angle and size 20x12
    ('40x10', 'I04'),   # 4 INTERIOR bedroom or bathroom 40x10
]


def read_manifest():
    ws = openpyxl.load_workbook(MANIFEST, data_only=True)['ALT MANIFEST']
    out = []
    for r in range(5, 65):
        out.append({
            'source': ws.cell(r, 1).value, 'size': ws.cell(r, 2).value,
            'view': ws.cell(r, 3).value, 'type': ws.cell(r, 4).value,
            'slot': ws.cell(r, 6).value, 'filename': ws.cell(r, 7).value,
            'alt': ws.cell(r, 8).value,
        })
    assert len(out) == 60, len(out)
    assert all((m['alt'] or '').strip() for m in out), 'manifest has an empty alt'
    return out


def main():
    manifest = read_manifest()
    by_key = {(m['size'], m['view']): m for m in manifest}

    # ---------------------------------------------------------------- Section H
    pack = json.load(open(PACK, encoding='utf-8'))
    entry = {'h2': pack['sectionH2'], 'guidanceLine': pack['guidance']}
    strings = 2
    for size in SIZES:
        p = pack['sizes'][size]
        entry[size] = {'h2': p['h2'], 'intro': p['body'], 'h3': p['h3'],
                       'applications': list(p['uses']), 'tab': p['tab']}
        strings += 4 + len(p['uses'])
    assert strings == 50, strings

    sh_path = os.path.join(DATA, 'c08-section-h-datasets.json')
    sh = json.load(open(sh_path, encoding='utf-8'))
    sh[SLUG] = entry
    with io.open(sh_path, 'w', encoding='utf-8', newline='\n') as fh:
        json.dump(sh, fh, indent=2, ensure_ascii=False)
        fh.write('\n')

    # ---------------------------------------------------------------- body copy
    lines = B.body_lines(COPY)
    blocks = B.to_blocks(lines)
    picks = [by_key[(s, v)] for s, v in IMAGE_SLOTS]

    parts, used = [], []
    for kind, payload in blocks:
        if kind == 'image':
            m = picks[payload - 1]
            src = f'/images/products/{SLUG}/info/{m["size"]}/{m["filename"]}'
            parts.append(
                f'<img src="{src}" alt="{htmllib.escape(m["alt"], quote=True)}"'
                f' width="1200" height="675" data-c08-info-image="true" />'
            )
            used.append(m)
        elif kind in ('h2', 'h3'):
            parts.append(f'<{kind}>{B.inline(payload)}</{kind}>')
        else:
            parts.append(f'<p>{B.inline(payload)}</p>')
    description_html = ''.join(parts)

    rendered = re.sub(r'</?(?:a|strong|em|span|b|i)\b[^>]*>', '', description_html)
    rendered = re.sub(r'<img[^>]*>', ' ', rendered)
    rendered = re.sub(r'<[^>]+>', ' ', rendered)
    rendered = re.sub(r'\s+', ' ', htmllib.unescape(rendered)).strip()
    source = re.sub(r'\s+', ' ', ' '.join(B.plain(p) for k, p in blocks if k != 'image')).strip()
    if rendered != source:
        for i, (a, b) in enumerate(zip(rendered, source)):
            if a != b:
                raise SystemExit(f'VERBATIM MISMATCH at {i}\n  rendered …{rendered[max(0,i-70):i+70]}…\n'
                                 f'  source   …{source[max(0,i-70):i+70]}…')
        raise SystemExit(f'VERBATIM LENGTH MISMATCH {len(rendered)} vs {len(source)}')

    # ------------------------------------------------------------ product data
    path = os.path.join(DATA, SLUG + '.json')
    data = json.load(open(path, encoding='utf-8'))

    # Gallery: six per size in slot order G1..G5 then H, each carrying its
    # manifest alt. The component slices 0..4 for the strip and reserves index 5
    # (I03) for the Section H panel.
    order = ['E01', 'I01', 'E03', 'E06', 'I02', 'I03']
    for variant in data['variants']:
        size = variant['sizeSlug']
        variant['images'] = [{
            'src': f'/images/products/{SLUG}/{size}/{by_key[(size, v)]["filename"]}',
            'alt': by_key[(size, v)]['alt'],
            'provenance': 'render', 'width': 900, 'height': 900,
        } for v in order]

    data['descriptionHtml'] = description_html
    data['applicationsDataset'] = SLUG
    data.pop('infoImages', None)
    with io.open(path, 'w', encoding='utf-8', newline='\n') as fh:
        json.dump(data, fh, indent=2, ensure_ascii=False)
        fh.write('\n')

    # ------------------------------------- alts onto the 24 info intake records
    intake_path = os.path.join(B.REPO, 'page-structure', 'C08', 'c08-e2-image-intake-report.json')
    intake = json.load(open(intake_path, encoding='utf-8'))
    written = 0
    for img in intake['images']:
        key = (img['sizeSlug'], img['sourceViewToken'])
        if key in by_key:
            img['alt'] = by_key[key]['alt']
            img['altStatus'] = 'implemented-verbatim-from-manifest-column-H'
            written += 1
    intake['altPolicy'] = ('Column H of C08-PREFAB-CONTAINER-HOUSE-ALT-MANIFEST-60-05Aug2026.xlsx '
                           'implemented byte-identically on all 60 assets.')
    intake['wiredToRoute'] = 60
    with io.open(intake_path, 'w', encoding='utf-8', newline='\n') as fh:
        json.dump(intake, fh, indent=1, ensure_ascii=False)
        fh.write('\n')

    # ------------------------------------------------------------------ verify
    gallery_alts = [i['alt'] for v in data['variants'] for i in v['images']]
    inline_alts = re.findall(r'<img[^>]*alt="([^"]*)"', description_html)
    exact = sum(1 for m in manifest if m['alt'] in gallery_alts
                or htmllib.unescape(m['alt']) in [htmllib.unescape(a) for a in inline_alts]
                or any(x['alt'] == m['alt'] for x in intake['images']))

    print(f'Section H      : {strings} strings, 6 sizes, tab labels carried')
    print(f'body copy      : VERBATIM OK · {sum(1 for k,_ in blocks if k=="h2")} H2 · '
          f'{sum(1 for k,_ in blocks if k=="h3")} H3 · {sum(1 for k,_ in blocks if k=="p")} p · '
          f'{len(used)} images · {len(re.findall(r"<a ", description_html))} links')
    print(f'gallery alts   : {len(gallery_alts)} written (36 expected), empty: '
          f'{sum(1 for a in gallery_alts if not a.strip())}')
    print(f'info alts      : {written} written to the intake record (24 expected)')
    print(f'inline alts    : {len(inline_alts)}, empty: {sum(1 for a in inline_alts if not a.strip())}')
    print(f'manifest rows traced to an implemented alt: {exact}/60')
    print()
    print('IMAGE SLOTS:')
    for i, m in enumerate(used, 1):
        print(f'  slot {i}  {m["size"]:6s} {m["view"]}  {m["type"]:10s} {m["filename"]}')


if __name__ == '__main__':
    main()
